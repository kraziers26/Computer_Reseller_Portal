from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required
from ..auth_utils import require_role
from ..db import db_cursor

admin_bp = Blueprint('admin', __name__)


@admin_bp.route('/dashboard')
@login_required
@require_role('admin')
def dashboard():
    # Filters
    f_month       = request.args.get('month', '')
    f_year        = request.args.get('year', '')
    f_company     = request.args.get('company', type=int)
    f_retailer    = request.args.get('retailer', '')
    f_submitter   = request.args.get('submitter', type=int)
    f_person      = request.args.get('person_by', type=int)
    f_card        = request.args.get('card', '')
    f_duplicates  = request.args.get('duplicates', '')
    f_order       = request.args.get('order_number', '')
    f_role        = request.args.get('role', '')
    f_needs_review = request.args.get('needs_review', '')
    f_fulfillment  = request.args.get('fulfillment', '')
    f_stuck_days   = request.args.get('stuck_days', type=int)
    f_batch        = request.args.get('batch', '')

    conditions = ["t.price_total > 0", "t.is_active = TRUE"]
    params = []
    if not f_duplicates:
        conditions.append("t.is_duplicate = FALSE")
    else:
        conditions.append("t.is_duplicate = TRUE")
    if f_month:
        conditions.append("TO_CHAR(t.purchase_date,'MM') = %s"); params.append(f_month)
    if f_year:
        conditions.append("TO_CHAR(t.purchase_date,'YYYY') = %s"); params.append(f_year)
    if f_company:
        conditions.append("t.company_id = %s"); params.append(f_company)
    if f_retailer:
        conditions.append("t.retailer = %s"); params.append(f_retailer)
    if f_submitter:
        conditions.append("t.submitted_by_user_id = %s"); params.append(f_submitter)
    if f_person:
        conditions.append("t.user_id = %s"); params.append(f_person)
    if f_card:
        conditions.append("t.card_id = %s"); params.append(f_card)
    if f_order:
        conditions.append("t.order_number ILIKE %s"); params.append(f'%{f_order}%')
    if f_fulfillment:
        conditions.append("t.fulfillment_status = %s"); params.append(f_fulfillment)
    if f_stuck_days:
        conditions.append(
            "EXTRACT(EPOCH FROM (NOW() - COALESCE(t.fulfillment_status_updated_at, t.submitted_at)))"
            " / 86400 >= %s")
        params.append(f_stuck_days)
    if f_role == 'contributor':
        conditions.append("sub.portal_role = 'contributor'")
    elif f_role == 'admin':
        conditions.append("sub.portal_role = 'admin'")
    if f_needs_review:
        conditions.append("t.review_status = 'Needs Review'")
    if f_batch:
        conditions.append("t.print_batch_id ILIKE %s"); params.append(f'%{f_batch}%')

    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT
                COUNT(*)                                              AS total_orders,
                ROUND(SUM(t.price_total)::numeric, 2)                AS total_gmv,
                ROUND(SUM(COALESCE(t.gross_paid_amount,0))::numeric,2) AS total_gross_paid,
                ROUND(SUM(COALESCE(t.net_paid_amount,0))::numeric,2)   AS total_net_paid,
                ROUND(SUM(COALESCE(t.sales_payroll_tax_withheld,0))::numeric,2) AS total_tax,
                ROUND(SUM(COALESCE(t.cashback_value,0))::numeric,2)    AS total_cashback,
                COUNT(*) FILTER (WHERE t.review_status='Pending')        AS pending_count,
                COUNT(*) FILTER (WHERE t.is_duplicate=TRUE)              AS dup_count,
                ROUND(SUM(COALESCE(t.costco_taxes_paid,0))::numeric,2) AS total_costco_taxes
            FROM transactions t {where}
        """, params)
        metrics = cur.fetchone()
        
        # Needs Review count — query separately without is_duplicate filter
        cur.execute("""
            SELECT COUNT(*) AS n FROM transactions
            WHERE is_active=TRUE AND review_status='Needs Review'
        """)
        needs_review_count = cur.fetchone()['n']

        # Recent submissions — join sub for role filter
        cur.execute(f"""
            SELECT t.order_number, t.retailer, t.purchase_date,
                   t.price_total, t.costco_taxes_paid, t.review_status, t.submitted_at,
                   t.is_duplicate, sub.username AS submitter_name, sub.portal_role AS submitter_role,
                   per.username AS person_name,
                   c.company_name, t.card_id,
                   d.cashback_rate
            FROM transactions t
            LEFT JOIN dim_users sub    ON t.submitted_by_email = sub.email
            LEFT JOIN dim_users per    ON t.user_id    = per.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            LEFT JOIN dim_cards d      ON t.card_id    = d.card_id
            {where}
            ORDER BY t.submitted_at DESC LIMIT 25
        """, params)
        recent = cur.fetchall()

        cur.execute("SELECT COUNT(*) AS n FROM v_pending_review")
        pending_total = cur.fetchone()['n']

        # Pipeline funnel counts + timing
        cur.execute("""
            SELECT
                COUNT(*) FILTER (WHERE fulfillment_status='uploaded' AND is_active=TRUE
                                   AND is_duplicate=FALSE)                        AS uploaded,
                COUNT(*) FILTER (WHERE fulfillment_status='batched'  AND is_active=TRUE)  AS batched,
                COUNT(*) FILTER (WHERE fulfillment_status='pending'  AND is_active=TRUE)  AS pending,
                COUNT(*) FILTER (WHERE fulfillment_status='received' AND is_active=TRUE)  AS received,
                COUNT(*) FILTER (WHERE fulfillment_status='invoiced' AND is_active=TRUE)  AS invoiced,
                COUNT(*) FILTER (WHERE review_status='Needs Review'  AND is_active=TRUE)  AS needs_review_total,
                COUNT(*) FILTER (WHERE review_status='Duplicate'     AND is_active=TRUE)  AS duplicates_total,
                ROUND(AVG(EXTRACT(EPOCH FROM (NOW()-COALESCE(fulfillment_status_updated_at,submitted_at)))/86400)
                    FILTER (WHERE fulfillment_status='uploaded' AND is_active=TRUE
                              AND is_duplicate=FALSE))                            AS avg_days_uploaded,
                ROUND(AVG(EXTRACT(EPOCH FROM (NOW()-COALESCE(fulfillment_status_updated_at,submitted_at)))/86400)
                    FILTER (WHERE fulfillment_status='batched'  AND is_active=TRUE)) AS avg_days_batched,
                ROUND(MAX(EXTRACT(EPOCH FROM (NOW()-COALESCE(fulfillment_status_updated_at,submitted_at)))/86400)
                    FILTER (WHERE fulfillment_status='batched'  AND is_active=TRUE)) AS max_days_batched,
                COUNT(*) FILTER (WHERE fulfillment_status='batched' AND is_active=TRUE
                    AND EXTRACT(EPOCH FROM (NOW()-COALESCE(fulfillment_status_updated_at,submitted_at)))/86400 >= 14
                ) AS stuck_batched
            FROM transactions
        """)
        pipeline = cur.fetchone()

        # Filter options
        cur.execute("SELECT DISTINCT retailer FROM transactions WHERE is_active=TRUE ORDER BY retailer")
        retailers = [r['retailer'] for r in cur.fetchall()]
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()
        cur.execute("""
            SELECT d.card_id, d.cashback_rate FROM dim_cards d
            WHERE d.card_id IN (SELECT DISTINCT card_id FROM transactions WHERE card_id IS NOT NULL AND is_active=TRUE)
            ORDER BY d.card_id
        """)
        cards = cur.fetchall()

        # Years available
        cur.execute("SELECT DISTINCT TO_CHAR(purchase_date,'YYYY') AS yr FROM transactions WHERE purchase_date IS NOT NULL ORDER BY yr DESC")
        years = [r['yr'] for r in cur.fetchall()]

    return render_template('dashboard.html',
                           metrics=metrics, recent=recent, pending_total=pending_total,
                           needs_review_count=needs_review_count,
                           pipeline=pipeline,
                           retailers=retailers, companies=companies, users=users, cards=cards, years=years,
                           filters={'month':f_month,'year':f_year,'company':f_company,
                                    'retailer':f_retailer,'submitter':f_submitter,
                                    'person_by':f_person,'card':f_card,
                                    'duplicates':f_duplicates,'order_number':f_order,'batch':f_batch,
                                    'role':f_role,'needs_review':f_needs_review,
                                    'fulfillment':f_fulfillment,'stuck_days':f_stuck_days})




@admin_bp.route('/duplicate-cleanup', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def duplicate_cleanup():
    from ..security import audit

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'delete_all':
            with db_cursor() as (cur, conn):
                # Find all duplicate tids (keep first per order_number)
                cur.execute("""
                    SELECT ARRAY_AGG(transaction_id ORDER BY submitted_at) AS tids
                    FROM transactions
                    WHERE is_active=TRUE AND order_number IS NOT NULL AND order_number != ''
                      AND (order_type IS NULL OR order_type NOT ILIKE '%return%')
                    GROUP BY order_number
                    HAVING COUNT(*) > 1
                """)
                rows = cur.fetchall()
                to_delete = []
                for row in rows:
                    to_delete.extend(str(t) for t in row['tids'][1:])
                if to_delete:
                    cur.execute("DELETE FROM receiving_item_lines WHERE transaction_item_id IN (SELECT item_id FROM transaction_items WHERE transaction_id = ANY(%s::uuid[]))", (to_delete,))
                    cur.execute("DELETE FROM receiving_items WHERE transaction_id = ANY(%s::uuid[])", (to_delete,))
                    cur.execute("DELETE FROM invoice_items WHERE transaction_id = ANY(%s::uuid[])", (to_delete,))
                    cur.execute("DELETE FROM transaction_items WHERE transaction_id = ANY(%s::uuid[])", (to_delete,))
                    cur.execute("DELETE FROM transactions WHERE transaction_id = ANY(%s::uuid[])", (to_delete,))
            flash(f'Deleted {len(to_delete)} duplicate transactions.', 'success')
            for tid in to_delete:
                audit('bulk_delete_duplicate', 'transaction', tid)
            return redirect(url_for('admin.duplicate_cleanup'))

        elif action == 'delete_selected':
            tids = request.form.getlist('tids')
            if tids:
                with db_cursor() as (cur, conn):
                    cur.execute("DELETE FROM receiving_item_lines WHERE transaction_item_id IN (SELECT item_id FROM transaction_items WHERE transaction_id = ANY(%s::uuid[]))", (tids,))
                    cur.execute("DELETE FROM receiving_items WHERE transaction_id = ANY(%s::uuid[])", (tids,))
                    cur.execute("DELETE FROM invoice_items WHERE transaction_id = ANY(%s::uuid[])", (tids,))
                    cur.execute("DELETE FROM transaction_items WHERE transaction_id = ANY(%s::uuid[])", (tids,))
                    cur.execute("DELETE FROM transactions WHERE transaction_id = ANY(%s::uuid[])", (tids,))
                flash(f'Deleted {len(tids)} selected duplicate transactions.', 'success')
                for tid in tids:
                    audit('delete_duplicate', 'transaction', tid)
            return redirect(url_for('admin.duplicate_cleanup'))

    # Build groups of duplicates
    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT
                t.order_number,
                MIN(t.retailer) AS retailer,
                ARRAY_AGG(t.transaction_id ORDER BY t.submitted_at) AS tids,
                ARRAY_AGG(COALESCE(u.username, t.submitted_by_email) ORDER BY t.submitted_at) AS submitters,
                ARRAY_AGG(t.submitted_at::date::text ORDER BY t.submitted_at) AS dates,
                ARRAY_AGG(t.price_total ORDER BY t.submitted_at) AS totals
            FROM transactions t
            LEFT JOIN dim_users u ON t.submitted_by_email = u.email
            WHERE t.is_active=TRUE AND t.order_number IS NOT NULL AND t.order_number != ''
              AND (t.order_type IS NULL OR t.order_type NOT ILIKE '%return%')
            GROUP BY t.order_number
            HAVING COUNT(*) > 1
            ORDER BY COUNT(*) DESC, MIN(t.submitted_at) DESC
        """)
        rows = cur.fetchall()

    groups = []
    total_dupes = 0
    for row in rows:
        tids       = [str(t) for t in row['tids']]
        submitters = row['submitters']
        dates      = row['dates']
        totals     = row['totals']
        kept_tid   = tids[0]
        dup_tids   = tids[1:]
        total_dupes += len(dup_tids)
        groups.append({
            'order_number':   row['order_number'],
            'retailer':       row['retailer'],
            'kept_tid':       kept_tid,
            'kept_submitter': submitters[0] if submitters else '—',
            'kept_date':      dates[0] if dates else '—',
            'dup_tids':       dup_tids,
            'dup_submitters': submitters[1:],
            'price_total':    totals[0] if totals else 0,
        })

    return render_template('duplicate_cleanup.html', groups=groups, total_dupes=total_dupes)

@admin_bp.route('/submissions/export')
@login_required
@require_role('admin')
def export_submissions():
    import io
    from openpyxl import Workbook
    from flask import send_file
    f_retailer   = request.args.get('retailer', '')
    f_company    = request.args.get('company', type=int)
    f_status     = request.args.get('status', '')
    f_month      = request.args.get('month', '')
    f_duplicates = request.args.get('duplicates', '')
    f_submitter  = request.args.get('submitter', type=int)
    f_card       = request.args.get('card', '')
    f_person     = request.args.get('person_by', type=int)
    f_order      = request.args.get('order_number', '')
    f_role       = request.args.get('role', '')
    f_fulfillment = request.args.get('fulfillment', '')
    f_stuck_days  = request.args.get('stuck_days', type=int)

    conditions = ["t.is_active = TRUE"]
    params = []
    if f_retailer: conditions.append("t.retailer = %s"); params.append(f_retailer)
    if f_company:  conditions.append("t.company_id = %s"); params.append(f_company)
    if f_status:   conditions.append("t.review_status = %s"); params.append(f_status)
    if f_month:    conditions.append("t.purchase_year_month = %s"); params.append(f_month)
    if f_duplicates: conditions.append("t.is_duplicate = TRUE")
    if f_submitter: conditions.append("sub.user_id = %s"); params.append(f_submitter)
    if f_card:     conditions.append("t.card_id = %s"); params.append(f_card)
    if f_person:   conditions.append("t.user_id = %s"); params.append(f_person)
    if f_order:    conditions.append("t.order_number ILIKE %s"); params.append(f'%{f_order}%')
    if f_role == 'contributor': conditions.append("sub.portal_role = 'contributor'")
    elif f_role == 'admin':     conditions.append("sub.portal_role = 'admin'")
    if f_fulfillment: conditions.append("t.fulfillment_status = %s"); params.append(f_fulfillment)
    if f_stuck_days:
        conditions.append("EXTRACT(EPOCH FROM (NOW()-COALESCE(t.fulfillment_status_updated_at,t.submitted_at)))/86400 >= %s")
        params.append(f_stuck_days)
    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT t.order_number, t.retailer, t.purchase_date, t.price_total,
                   t.review_status, t.fulfillment_status, t.submitted_at,
                   t.fulfillment_status_updated_at,
                   ROUND(EXTRACT(EPOCH FROM (NOW()-COALESCE(t.fulfillment_status_updated_at,
                         t.submitted_at)))/86400) AS days_in_status,
                   sub.username AS submitter, per.username AS person_by,
                   c.company_name, t.card_id, t.order_type, t.is_duplicate
            FROM transactions t
            LEFT JOIN dim_users sub    ON t.submitted_by_email = sub.email
            LEFT JOIN dim_users per    ON t.user_id = per.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            {where}
            ORDER BY t.submitted_at DESC
        """, params)
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    headers = ['Order #','Retailer','Purchase Date','Total','Review Status',
               'Fulfillment Stage','Days in Stage','Submitted At','Submitter',
               'Person By','Company','Card','Order Type','Duplicate']
    ws.append(headers)
    for r in rows:
        ws.append([
            r['order_number'], r['retailer'],
            r['purchase_date'].strftime('%Y-%m-%d') if r['purchase_date'] else '',
            float(r['price_total'] or 0), r['review_status'],
            r['fulfillment_status'], int(r['days_in_status'] or 0),
            r['submitted_at'].strftime('%Y-%m-%d %H:%M') if r['submitted_at'] else '',
            r['submitter'], r['person_by'], r['company_name'],
            r['card_id'], r['order_type'], 'Yes' if r['is_duplicate'] else 'No'
        ])
    from openpyxl.styles import Font
    for cell in ws[1]: cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='submissions_export.xlsx')


@admin_bp.route('/submissions/bulk-action', methods=['POST'])
@login_required
@require_role('admin')
def bulk_action():
    from ..security import audit
    action = request.form.get('action')
    tids   = request.form.getlist('tids')
    back   = request.form.get('back', url_for('admin.all_submissions'))
    if not tids:
        flash('No transactions selected.', 'error')
        return redirect(back)
    with db_cursor() as (cur, conn):
        if action == 'approve':
            cur.execute(
                "UPDATE transactions SET review_status='Reviewed', review_date=NOW() "
                "WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            flash(f'{len(tids)} transaction(s) approved.', 'success')
        elif action == 'flag':
            cur.execute(
                "UPDATE transactions SET review_status='Flagged' "
                "WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            flash(f'{len(tids)} transaction(s) flagged.', 'warning')
        elif action == 'mark_duplicate':
            cur.execute(
                "UPDATE transactions SET is_duplicate=TRUE, review_status='Duplicate' "
                "WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            flash(f'{len(tids)} transaction(s) marked as duplicate.', 'warning')
        elif action == 'delete':
            cur.execute("DELETE FROM receiving_item_lines WHERE transaction_item_id IN (SELECT item_id FROM transaction_items WHERE transaction_id = ANY(%s::uuid[]))", (tids,))
            cur.execute("DELETE FROM receiving_items WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            cur.execute("DELETE FROM invoice_items WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            cur.execute("DELETE FROM transaction_items WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            cur.execute("DELETE FROM transactions WHERE transaction_id = ANY(%s::uuid[])", (tids,))
            flash(f'{len(tids)} transaction(s) permanently deleted.', 'danger')
        else:
            flash('Unknown action.', 'error')
    for tid in tids:
        audit(f'bulk_{action}', 'transaction', tid)
    return redirect(back)


@admin_bp.route('/submissions/all')
@login_required
@require_role('admin')
def all_submissions():
    page     = request.args.get('page', 1, type=int)
    per_page = 25
    offset   = (page - 1) * per_page

    f_retailer   = request.args.get('retailer', '')
    f_company    = request.args.get('company', type=int)
    f_status     = request.args.get('status', '')
    f_month      = request.args.get('month', '')
    f_duplicates = request.args.get('duplicates', '')
    f_submitter  = request.args.get('submitter', type=int)
    f_card       = request.args.get('card', '')
    f_person     = request.args.get('person_by', type=int)
    f_order      = request.args.get('order_number', '')
    f_role        = request.args.get('role', '')
    f_fulfillment  = request.args.get('fulfillment', '')
    f_stuck_days   = request.args.get('stuck_days', type=int)
    f_batch        = request.args.get('batch', '')

    conditions = ["t.is_active = TRUE"]
    params = []
    if f_retailer:
        conditions.append("t.retailer = %s"); params.append(f_retailer)
    if f_company:
        conditions.append("t.company_id = %s"); params.append(f_company)
    if f_status:
        conditions.append("t.review_status = %s"); params.append(f_status)
    if f_role == 'contributor':
        conditions.append("sub.portal_role = 'contributor'")
    elif f_role == 'admin':
        conditions.append("sub.portal_role = 'admin'")
    if f_month:
        conditions.append("t.purchase_year_month = %s"); params.append(f_month)
    if f_duplicates:
        conditions.append("t.is_duplicate = TRUE")
    if f_submitter:
        conditions.append("sub.user_id = %s"); params.append(f_submitter)
    if f_card:
        conditions.append("t.card_id = %s"); params.append(f_card)
    if f_person:
        conditions.append("t.user_id = %s"); params.append(f_person)
    if f_order:
        conditions.append("t.order_number ILIKE %s"); params.append(f'%{f_order}%')
    if f_fulfillment:
        conditions.append("t.fulfillment_status = %s"); params.append(f_fulfillment)
    if f_stuck_days:
        conditions.append(
            "EXTRACT(EPOCH FROM (NOW() - COALESCE(t.fulfillment_status_updated_at, t.submitted_at)))"
            " / 86400 >= %s")
        params.append(f_stuck_days)

    if f_batch:
        conditions.append("t.print_batch_id = %s"); params.append(f_batch)

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT t.transaction_id, t.order_number, t.retailer,
                   t.purchase_date, t.price_total, t.order_type,
                   t.review_status, t.submitted_at, t.is_duplicate,
                   t.card_id, d.cashback_rate,
                   ROUND(COALESCE(t.gross_paid_amount,0)::numeric,2) AS gross_paid,
                   ROUND(COALESCE(t.net_paid_amount,0)::numeric,2)   AS net_paid,
                   ROUND(COALESCE(t.sales_payroll_tax_withheld,0)::numeric,2) AS tax_withheld,
                   ROUND(COALESCE(t.cashback_value,0)::numeric,2)    AS cashback,
                   sub.username AS submitter_name,
                   per.username AS person_name,
                   c.company_name, t.notes
            FROM transactions t
            LEFT JOIN dim_users sub    ON t.submitted_by_email = sub.email
            LEFT JOIN dim_users per    ON t.user_id    = per.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            LEFT JOIN dim_cards d      ON t.card_id    = d.card_id
            {where}
            ORDER BY t.submitted_at DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        submissions = cur.fetchall()

        cur.execute(f"""
            SELECT COUNT(*) AS n FROM transactions t
            LEFT JOIN dim_users sub ON t.submitted_by_email = sub.email
            {where}
        """, params)
        total = cur.fetchone()['n']

        cur.execute("SELECT DISTINCT retailer FROM transactions WHERE is_active=TRUE ORDER BY retailer")
        retailers = [r['retailer'] for r in cur.fetchall()]
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE")
        companies = cur.fetchall()
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()
        cur.execute("""
            SELECT d.card_id, d.cashback_rate FROM dim_cards d
            WHERE d.card_id IN (SELECT DISTINCT card_id FROM transactions WHERE card_id IS NOT NULL AND is_active=TRUE)
            ORDER BY d.card_id
        """)
        cards = cur.fetchall()
        cur.execute("SELECT DISTINCT purchase_year_month FROM transactions WHERE is_active=TRUE ORDER BY purchase_year_month DESC")
        months = [r['purchase_year_month'] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT print_batch_id FROM transactions WHERE print_batch_id IS NOT NULL AND is_active=TRUE ORDER BY print_batch_id")
        batches = [r['print_batch_id'] for r in cur.fetchall()]

    return render_template('all_submissions.html',
                           submissions=submissions, total=total,
                           page=page, per_page=per_page,
                           retailers=retailers, companies=companies, users=users, cards=cards, months=months,
                           batches=batches,
                           filters={'retailer':f_retailer,'company':f_company,'status':f_status,
                                    'month':f_month,'duplicates':f_duplicates,'submitter':f_submitter,
                                    'card':f_card,'person_by':f_person,'order_number':f_order,
                                    'role':f_role,'fulfillment':f_fulfillment,
                                    'stuck_days':f_stuck_days,'batch':f_batch})


@admin_bp.route('/submissions/<uuid:tid>', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def review_submission(tid):
    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT t.*, sub.username AS submitter_name, per.username AS person_name,
                   c.company_name,
                   (t.invoice_pdf IS NOT NULL) AS has_pdf_in_db,
                   t.membership_number
            FROM transactions t
            LEFT JOIN dim_users sub    ON t.submitted_by_email = sub.email
            LEFT JOIN dim_users per    ON t.user_id    = per.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            WHERE t.transaction_id = %s
        """, (str(tid),))
        txn = cur.fetchone()
        if not txn:
            flash('Transaction not found.', 'error')
            return redirect(url_for('admin.all_submissions'))
        cur.execute("SELECT * FROM transaction_items WHERE transaction_id=%s", (str(tid),))
        items = cur.fetchall()
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE")
        companies = cur.fetchall()
        cur.execute("SELECT card_id, card_name, card_brand, cashback_rate FROM dim_cards WHERE is_active=TRUE ORDER BY card_id")
        cards = cur.fetchall()
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()

    if request.method == 'POST':
        action = request.form.get('action')
        with db_cursor() as (cur, conn):
            from ..security import audit
            if action == 'approve':
                cur.execute("UPDATE transactions SET review_status='Reviewed', review_date=NOW() WHERE transaction_id=%s", (str(tid),))
                audit('transaction_approved', 'transaction', str(tid))
                flash('Transaction approved.', 'success')
            elif action == 'flag':
                cur.execute("UPDATE transactions SET review_status='Flagged' WHERE transaction_id=%s", (str(tid),))
                audit('transaction_flagged', 'transaction', str(tid))
                flash('Transaction flagged.', 'warning')
            elif action == 'mark_duplicate':
                cur.execute("UPDATE transactions SET is_duplicate=TRUE, review_status='Flagged' WHERE transaction_id=%s", (str(tid),))
                audit('transaction_marked_duplicate', 'transaction', str(tid))
                flash('Marked as duplicate.', 'warning')
            elif action == 'inactivate':
                cur.execute("UPDATE transactions SET is_active=FALSE WHERE transaction_id=%s", (str(tid),))
                audit('transaction_inactivated', 'transaction', str(tid))
                flash('Transaction inactivated. It will no longer appear in reports.', 'warning')
                return redirect(url_for('admin.all_submissions'))
            elif action == 'delete':
                cur.execute("DELETE FROM receiving_item_lines WHERE transaction_item_id IN (SELECT item_id FROM transaction_items WHERE transaction_id=%s)", (str(tid),))
                cur.execute("DELETE FROM receiving_items WHERE transaction_id=%s", (str(tid),))
                cur.execute("DELETE FROM invoice_items WHERE transaction_id=%s", (str(tid),))
                cur.execute("DELETE FROM transaction_items WHERE transaction_id=%s", (str(tid),))
                cur.execute("DELETE FROM transactions WHERE transaction_id=%s", (str(tid),))
                audit('transaction_deleted', 'transaction', str(tid))
                flash('Transaction permanently deleted.', 'danger')
                return redirect(url_for('admin.all_submissions'))
            elif action == 'edit':
                import uuid as _uuid
                card_id           = request.form.get('card_id') or None
                company_id        = request.form.get('company_id', type=int)
                user_id           = request.form.get('user_id', type=int)
                order_type        = request.form.get('order_type')
                retailer          = request.form.get('retailer', '').strip()
                order_number      = request.form.get('order_number', '').strip()
                purchase_date     = request.form.get('purchase_date') or None
                price             = request.form.get('price_total', type=float)
                costco_taxes      = request.form.get('costco_taxes', type=float) or None
                fulfillment_st    = request.form.get('fulfillment_status') or None
                review_st         = request.form.get('review_status') or 'Reviewed'
                notes             = request.form.get('notes', '').strip()[:140] or None
                membership_number = request.form.get('membership_number', '').strip() or None

                cashback_rate = cashback_value = None
                if card_id and price:
                    cur.execute("SELECT cashback_rate FROM dim_cards WHERE card_id=%s", (card_id,))
                    crow = cur.fetchone()
                    if crow:
                        cashback_rate  = float(crow['cashback_rate'])
                        cashback_value = round(price * cashback_rate, 2)

                gross_paid   = round(price * 0.01, 2) if price else None
                net_paid     = round(gross_paid * 0.8, 2) if gross_paid else None
                tax_withheld = round(gross_paid * 0.2, 2) if gross_paid else None
                gross_biz    = round((gross_paid or 0)+(cashback_value or 0),2) if gross_paid else None
                net_biz      = round((gross_biz or 0)-(net_paid or 0),2) if gross_biz else None

                # Parse year_month from purchase_date
                purchase_ym = None
                if purchase_date:
                    try:
                        from datetime import datetime as _dt
                        purchase_ym = _dt.strptime(purchase_date, '%Y-%m-%d').strftime('%Y-%m')
                    except Exception:
                        pass

                upd_fields = """
                    card_id=%s, company_id=%s, user_id=%s, order_type=%s,
                    retailer=%s, order_number=%s, purchase_date=%s, purchase_year_month=%s,
                    price_total=%s, costco_taxes_paid=%s,
                    cashback_rate=%s, cashback_value=%s,
                    gross_paid_amount=%s, net_paid_amount=%s,
                    gross_business_commission=%s, net_business_commission=%s,
                    sales_payroll_tax_withheld=%s, notes=%s, membership_number=%s,
                    review_status=%s, review_date=NOW()
                """
                upd_args = (card_id, company_id, user_id, order_type,
                            retailer, order_number, purchase_date, purchase_ym,
                            price, costco_taxes,
                            cashback_rate, cashback_value,
                            gross_paid, net_paid, gross_biz, net_biz, tax_withheld,
                            notes, membership_number, review_st)
                if fulfillment_st:
                    cur.execute(f"UPDATE transactions SET {upd_fields}, fulfillment_status=%s, fulfillment_status_updated_at=NOW() WHERE transaction_id=%s",
                                upd_args + (fulfillment_st, str(tid)))
                else:
                    cur.execute(f"UPDATE transactions SET {upd_fields} WHERE transaction_id=%s",
                                upd_args + (str(tid),))

                # Rebuild line items from form list fields
                descs = request.form.getlist('item_description[]')
                if descs:
                    cur.execute("DELETE FROM receiving_item_lines WHERE transaction_item_id IN (SELECT item_id FROM transaction_items WHERE transaction_id=%s)", (str(tid),))
                    cur.execute("DELETE FROM transaction_items WHERE transaction_id=%s", (str(tid),))
                    skus        = request.form.getlist('item_sku[]')
                    qtys        = request.form.getlist('item_qty[]')
                    unit_prices = request.form.getlist('item_unit_price[]')
                    line_totals = request.form.getlist('item_line_total[]')
                    for i, desc in enumerate(descs):
                        desc = desc.strip()
                        if not desc:
                            continue
                        try:
                            unit_p = float(unit_prices[i]) if i < len(unit_prices) else 0.0
                            line_t = float(line_totals[i]) if i < len(line_totals) else 0.0
                            qty    = int(qtys[i]) if i < len(qtys) else 1
                            sku    = (skus[i].strip() if i < len(skus) else '') or None
                        except (ValueError, IndexError):
                            continue
                        cur.execute("""
                            INSERT INTO transaction_items
                            (item_id, transaction_id, item_description, sku_model_color,
                             quantity, unit_price, line_total)
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                        """, (str(_uuid.uuid4()), str(tid), desc, sku, qty, unit_p, line_t))

                audit('transaction_edited', 'transaction', str(tid))
                flash('Transaction updated.', 'success')
        return redirect(url_for('admin.review_submission', tid=tid))

    return render_template('review_submission.html',
                           txn=txn, items=items, companies=companies, cards=cards, users=users)


@admin_bp.route('/payroll')
@login_required
@require_role('admin')
def payroll():
    month   = request.args.get('month', '')
    company = request.args.get('company', type=int)
    sort_by = request.args.get('sort', 'username')
    sort_dir = request.args.get('dir', 'asc')

    valid_sorts = {'username','order_count','total_purchases','gross_paid','net_paid','tax_withheld'}
    if sort_by not in valid_sorts:
        sort_by = 'username'
    order_clause = f"{sort_by} {'DESC' if sort_dir=='desc' else 'ASC'}"

    conditions = ["t.review_status != 'Flagged'", "t.is_duplicate = FALSE",
                  "t.price_total > 0", "t.is_active = TRUE"]
    params = []
    if month:
        conditions.append("t.purchase_year_month = %s"); params.append(month)

    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        # Separate queries per company
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()

        company_data = {}
        for comp in companies:
            if company and comp['company_id'] != company:
                continue
            cparams = params + [comp['company_id']]
            cur.execute(f"""
                SELECT
                    u.user_id, u.username,
                    t.purchase_year_month,
                    COUNT(t.transaction_id)                                    AS order_count,
                    ROUND(SUM(t.price_total)::numeric, 2)                      AS total_purchases,
                    ROUND(SUM(COALESCE(t.gross_paid_amount,0))::numeric,2)     AS gross_paid,
                    ROUND(SUM(COALESCE(t.net_paid_amount,0))::numeric,2)       AS net_paid,
                    ROUND(SUM(COALESCE(t.sales_payroll_tax_withheld,0))::numeric,2) AS tax_withheld
                FROM transactions t
                LEFT JOIN dim_users u ON t.user_id = u.user_id
                {where} AND t.company_id = %s
                GROUP BY u.user_id, u.username, t.purchase_year_month
                ORDER BY t.purchase_year_month DESC, {order_clause}
            """, cparams)
            company_data[comp['company_name']] = cur.fetchall()

        cur.execute("""
            SELECT DISTINCT purchase_year_month FROM transactions
            WHERE price_total > 0 AND is_active=TRUE ORDER BY purchase_year_month DESC
        """)
        months = [r['purchase_year_month'] for r in cur.fetchall()]

    return render_template('payroll.html',
                           company_data=company_data, months=months, companies=companies,
                           selected_month=month, selected_company=company,
                           sort_by=sort_by, sort_dir=sort_dir)


@admin_bp.route('/cashback')
@login_required
@require_role('admin')
def cashback():
    f_month   = request.args.get('month', '')
    f_year    = request.args.get('year', '')
    f_company = request.args.get('company', type=int)
    f_person  = request.args.get('person_by', type=int)

    conditions = ["t.is_active = TRUE"]
    t_params = []
    if f_month:
        conditions.append("TO_CHAR(t.purchase_date,'MM') = %s"); t_params.append(f_month)
    if f_year:
        conditions.append("TO_CHAR(t.purchase_date,'YYYY') = %s"); t_params.append(f_year)
    if f_person:
        conditions.append("t.user_id = %s"); t_params.append(f_person)
    if f_company:
        conditions.append("t.company_id = %s"); t_params.append(f_company)

    t_where = ('AND ' + ' AND '.join(conditions)) if conditions else ''

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT
                d.card_id, d.card_name, d.card_brand,
                u.username AS cardholder,
                c.company_name,
                d.cashback_rate,
                COUNT(t.transaction_id)                               AS transactions,
                ROUND(SUM(COALESCE(t.price_total,0))::numeric, 2)    AS total_spend,
                ROUND(SUM(COALESCE(t.cashback_value,0))::numeric, 2) AS total_cashback
            FROM dim_cards d
            LEFT JOIN transactions t  ON t.card_id    = d.card_id
                                      AND t.price_total > 0
                                      AND t.is_duplicate = FALSE
                                      {t_where}
            LEFT JOIN dim_users u     ON d.user_id    = u.user_id
            LEFT JOIN dim_companies c ON d.company_id = c.company_id
            WHERE d.is_active = TRUE
            GROUP BY d.card_id, d.card_name, d.card_brand,
                     u.username, c.company_name, d.cashback_rate
            ORDER BY total_cashback DESC NULLS LAST
        """, t_params)
        cashback_data = cur.fetchall()

        # Total cashback by company
        cur.execute(f"""
            SELECT c.company_name,
                   ROUND(SUM(COALESCE(t.cashback_value,0))::numeric,2) AS total
            FROM transactions t
            JOIN dim_companies c ON t.company_id = c.company_id
            WHERE t.price_total > 0 AND t.is_duplicate = FALSE {t_where}
            GROUP BY c.company_name ORDER BY total DESC
        """, t_params)
        company_cashback = cur.fetchall()

        cur.execute("SELECT DISTINCT TO_CHAR(purchase_date,'YYYY') AS yr FROM transactions WHERE purchase_date IS NOT NULL ORDER BY yr DESC")
        years = [r['yr'] for r in cur.fetchall()]
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()

    return render_template('cashback.html',
                           cashback_data=cashback_data, company_cashback=company_cashback,
                           years=years, companies=companies, users=users,
                           filters={'month':f_month,'year':f_year,'company':f_company,'person_by':f_person})


@admin_bp.route('/print-batch', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def print_batch():
    from flask_login import current_user
    if request.method == 'POST':
        txn_ids  = request.form.getlist('txn_ids')
        batch_id = request.form.get('batch_id', '').strip()
        skip_tids = request.form.getlist('skip_tids')
        if skip_tids:
            with db_cursor() as (cur, conn):
                cur.execute(
                    "UPDATE transactions SET skip_print=TRUE WHERE transaction_id = ANY(%s::uuid[])",
                    (skip_tids,))
            flash(f'{len(skip_tids)} invoice(s) moved to review pile.', 'info')
        if txn_ids and batch_id:
            with db_cursor() as (cur, conn):
                cur.execute("""
                    UPDATE transactions
                    SET print_batch_id=%s, print_date=NOW(),
                        fulfillment_status='batched',
                        fulfillment_status_updated_at=NOW()
                    WHERE transaction_id = ANY(%s::uuid[])
                """, (batch_id, txn_ids))
                # Give this batch a metadata row (name defaults to the ID) so it's
                # renameable right away. Safe if the batch_id already has one
                # (e.g. topping up an existing batch with more invoices).
                safe_name = _unique_batch_name(cur, batch_id, batch_id)
                cur.execute("""
                    INSERT INTO print_batches (batch_id, batch_name, created_by)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (batch_id) DO NOTHING
                """, (batch_id, safe_name, current_user.id if current_user.is_authenticated else None))
            flash(f'{len(txn_ids)} invoices tagged as batch {batch_id}.', 'success')

    f_retailer  = request.args.get('retailer', '')
    f_person    = request.args.get('person_by', type=int)
    f_submitter = request.args.get('submitter', type=int)
    f_company   = request.args.get('company', type=int)
    f_role      = request.args.get('role', '')
    f_date_from = request.args.get('date_from', '')
    f_date_to   = request.args.get('date_to', '')

    conditions = ["t.print_date IS NULL", "t.review_status != 'Flagged'",
                  "t.is_active=TRUE", "COALESCE(t.skip_print,FALSE)=FALSE"]
    params = []
    if f_retailer:  conditions.append("t.retailer=%s"); params.append(f_retailer)
    if f_person:    conditions.append("t.user_id=%s"); params.append(f_person)
    if f_submitter: conditions.append("sub.user_id=%s"); params.append(f_submitter)
    if f_company:   conditions.append("t.company_id=%s"); params.append(f_company)
    if f_role == 'contributor': conditions.append("sub.portal_role='contributor'")
    elif f_role == 'admin':     conditions.append("sub.portal_role='admin'")
    if f_date_from: conditions.append("t.purchase_date::date >= %s"); params.append(f_date_from)
    if f_date_to:   conditions.append("t.purchase_date::date <= %s"); params.append(f_date_to)
    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT t.transaction_id, t.order_number, t.retailer,
                   t.purchase_date, t.price_total, t.print_date,
                   t.print_batch_id, t.invoice_file_path,
                   (t.invoice_pdf IS NOT NULL) AS has_pdf,
                   t.skip_print,
                   u.username, c.company_name,
                   sub.portal_role AS submitter_role
            FROM transactions t
            LEFT JOIN dim_users u     ON t.user_id    = u.user_id
            LEFT JOIN dim_companies c ON t.company_id = c.company_id
            LEFT JOIN dim_users sub   ON t.submitted_by_email = sub.email
            {where}
            ORDER BY t.submitted_at DESC
        """, params)
        unprinted = cur.fetchall()
        cur.execute("""
            SELECT DISTINCT print_batch_id, MIN(print_date) AS batch_date, COUNT(*) AS cnt,
                   MIN(submitted_by_email) AS created_by_email
            FROM transactions WHERE print_batch_id IS NOT NULL
            GROUP BY print_batch_id ORDER BY batch_date DESC LIMIT 3
        """)
        batches = cur.fetchall()
        cur.execute("SELECT DISTINCT retailer FROM transactions WHERE is_active=TRUE ORDER BY retailer")
        retailers = [r['retailer'] for r in cur.fetchall()]
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()

    return render_template('print_batch.html', unprinted=unprinted, batches=batches,
                           retailers=retailers, users=users, companies=companies,
                           filters={'retailer':f_retailer,'person_by':f_person,'company':f_company,
                                    'submitter':f_submitter,'role':f_role,
                                    'date_from':f_date_from,'date_to':f_date_to})


@admin_bp.route('/batch-history')
@login_required
@require_role('admin')
def batch_history():
    f_batch     = request.args.get('batch_id', '').strip()
    f_company   = request.args.get('company', type=int)
    f_submitter = request.args.get('submitter', type=int)
    f_date_from = request.args.get('date_from', '')
    f_date_to   = request.args.get('date_to', '')
    f_order     = request.args.get('order_number', '').strip()

    conditions = ["t.print_batch_id IS NOT NULL"]
    params = []
    if f_batch:
        conditions.append("(t.print_batch_id ILIKE %s OR pb.batch_name ILIKE %s)")
        params.append(f'%{f_batch}%'); params.append(f'%{f_batch}%')
    if f_company:
        conditions.append("t.company_id = %s"); params.append(f_company)
    if f_submitter:
        conditions.append("u.user_id = %s"); params.append(f_submitter)
    if f_date_from:
        conditions.append("t.print_date::date >= %s"); params.append(f_date_from)
    if f_date_to:
        conditions.append("t.print_date::date <= %s"); params.append(f_date_to)
    if f_order:
        conditions.append("t.order_number ILIKE %s"); params.append(f'%{f_order}%')

    where = 'WHERE ' + ' AND '.join(conditions)

    # An order is still "editable" (removable / undo-able) only if it hasn't
    # progressed past Batched, AND hasn't already been scanned inside an open
    # receiving session (received/partial/missing there) even though the
    # transaction itself still says 'batched' until that session closes.
    editable_expr = """
        (t.fulfillment_status = 'batched'
         AND NOT EXISTS (
            SELECT 1 FROM receiving_items ri
            JOIN receiving_sessions rs ON ri.session_id = rs.session_id
            WHERE ri.transaction_id = t.transaction_id
              AND rs.status = 'open'
              AND ri.receive_status IN ('received','partial','missing')
         ))
    """

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT t.print_batch_id,
                   COALESCE(pb.batch_name, t.print_batch_id) AS batch_name,
                   MIN(t.print_date)         AS batch_date,
                   COUNT(*)                  AS cnt,
                   COUNT(*) FILTER (WHERE {editable_expr}) AS editable_count,
                   MIN(u.username)           AS created_by,
                   STRING_AGG(DISTINCT c.company_name, ', ') AS companies
            FROM transactions t
            LEFT JOIN dim_users u     ON t.submitted_by_email = u.email
            LEFT JOIN dim_companies c ON t.company_id = c.company_id
            LEFT JOIN print_batches pb ON pb.batch_id = t.print_batch_id
            {where}
            GROUP BY t.print_batch_id, pb.batch_name
            ORDER BY batch_date DESC
        """, params)
        batches = cur.fetchall()

        # When searching by order number, also fetch the individual matching orders
        order_results = []
        if f_order:
            cur.execute("""
                SELECT t.transaction_id, t.order_number, t.retailer,
                       t.purchase_date, t.price_total, t.print_batch_id,
                       t.fulfillment_status, t.print_date,
                       per.username AS person_name,
                       c.company_name
                FROM transactions t
                LEFT JOIN dim_users per    ON t.user_id    = per.user_id
                LEFT JOIN dim_companies c  ON t.company_id = c.company_id
                WHERE t.order_number ILIKE %s AND t.is_active=TRUE
                ORDER BY t.print_batch_id NULLS LAST, t.submitted_at DESC
            """, (f'%{f_order}%',))
            order_results = cur.fetchall()

        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()

    return render_template('batch_history.html', batches=batches, companies=companies,
                           users=users, order_results=order_results,
                           filters={'batch_id':f_batch,'company':f_company,
                                    'submitter':f_submitter,
                                    'date_from':f_date_from,'date_to':f_date_to,
                                    'order_number':f_order})


# An order is "editable" (removable from a batch / reverted by Undo) only if
# it hasn't progressed past Batched, AND hasn't already been scanned inside an
# open receiving session (received/partial/missing there) even though the
# transaction itself still reads 'batched' until that session is closed.
# An order is locked from editing if it currently has a non-returned line on
# any invoice. The moment it's marked Returned there (even after being
# invoiced), it's unlocked again -- this same rule is reused in Receiving.
_INVOICE_LOCKED_EXPR = """
    EXISTS (
        SELECT 1 FROM invoice_items ii
        WHERE ii.transaction_id = t.transaction_id AND NOT ii.returned
    )
"""

def _unique_batch_name(cur, batch_id, desired_name):
    """Return a batch_name guaranteed not to collide with a DIFFERENT batch's
    name (the unique index is case-insensitive). Appends ' (2)', ' (3)', etc.
    if the desired name is already taken by another batch_id. Without this,
    a plain INSERT can throw a UniqueViolation that rolls back the whole
    transaction -- including the transactions UPDATE that actually batches
    the orders, silently failing the entire print-batch action."""
    cur.execute(
        "SELECT 1 FROM print_batches WHERE LOWER(batch_name) = LOWER(%s) AND batch_id != %s",
        (desired_name, batch_id))
    if not cur.fetchone():
        return desired_name
    n = 2
    while True:
        candidate = f"{desired_name} ({n})"
        cur.execute(
            "SELECT 1 FROM print_batches WHERE LOWER(batch_name) = LOWER(%s) AND batch_id != %s",
            (candidate, batch_id))
        if not cur.fetchone():
            return candidate
        n += 1


_EDITABLE_EXPR = f"""
    (
        (t.exception_status = 'returned' AND NOT {_INVOICE_LOCKED_EXPR})
        OR
        (t.fulfillment_status = 'batched'
         AND NOT EXISTS (
            SELECT 1 FROM receiving_items ri
            JOIN receiving_sessions rs ON ri.session_id = rs.session_id
            WHERE ri.transaction_id = t.transaction_id
              AND rs.status = 'open'
              AND ri.receive_status IN ('received','partial','missing')
         ))
    )
"""

_LOCK_REASON_EXPR = f"""
    CASE
        WHEN t.exception_status = 'returned' AND NOT {_INVOICE_LOCKED_EXPR} THEN NULL
        WHEN {_INVOICE_LOCKED_EXPR} THEN 'On an active invoice — return it there first'
        WHEN t.fulfillment_status = 'invoiced' THEN 'Already invoiced'
        WHEN t.fulfillment_status = 'received' THEN 'Already received'
        WHEN EXISTS (
            SELECT 1 FROM receiving_items ri
            JOIN receiving_sessions rs ON ri.session_id = rs.session_id
            WHERE ri.transaction_id = t.transaction_id
              AND rs.status = 'open'
              AND ri.receive_status IN ('received','partial','missing')
        ) THEN 'Already counted in an open receiving session'
        ELSE NULL
    END
"""


@admin_bp.route('/batch/<batch_id>')
@login_required
@require_role('admin')
def batch_detail(batch_id):
    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT t.transaction_id, t.order_number, t.retailer,
                   t.purchase_date, t.price_total, t.invoice_file_path,
                   (t.invoice_pdf IS NOT NULL) AS has_pdf,
                   t.print_date, u.username AS person_name, c.company_name,
                   {_EDITABLE_EXPR} AS is_editable,
                   {_LOCK_REASON_EXPR} AS lock_reason
            FROM transactions t
            LEFT JOIN dim_users u     ON t.user_id    = u.user_id
            LEFT JOIN dim_companies c ON t.company_id = c.company_id
            WHERE t.print_batch_id = %s
            ORDER BY t.purchase_date
        """, (batch_id,))
        invoices = cur.fetchall()
        batch_date = invoices[0]['print_date'] if invoices else None

        cur.execute("SELECT batch_name FROM print_batches WHERE batch_id = %s", (batch_id,))
        pb = cur.fetchone()
        batch_name = pb['batch_name'] if pb else batch_id

        # Orders eligible to be pulled into this batch: currently unbatched, active.
        cur.execute("""
            SELECT t.transaction_id, t.order_number, t.retailer, t.purchase_date,
                   t.price_total, c.company_name
            FROM transactions t
            LEFT JOIN dim_companies c ON t.company_id = c.company_id
            WHERE t.fulfillment_status = 'uploaded' AND t.is_active = TRUE
            ORDER BY t.submitted_at DESC
            LIMIT 200
        """)
        addable_orders = cur.fetchall()

    return render_template('batch_detail.html', batch_id=batch_id, batch_name=batch_name,
                           invoices=invoices, batch_date=batch_date,
                           addable_orders=addable_orders)


@admin_bp.route('/batch/unbatch', methods=['POST'])
@login_required
@require_role('admin')
def unbatch():
    from ..security import audit

    batch_id = request.form.get('batch_id', '').strip()
    if not batch_id:
        return redirect(request.referrer or url_for('admin.batch_history'))

    with db_cursor() as (cur, conn):
        cur.execute(f"""
            SELECT t.transaction_id, t.order_number, {_EDITABLE_EXPR} AS is_editable,
                   {_LOCK_REASON_EXPR} AS lock_reason
            FROM transactions t
            WHERE t.print_batch_id = %s
        """, (batch_id,))
        rows = cur.fetchall()

        cur.execute("SELECT batch_name FROM print_batches WHERE batch_id = %s", (batch_id,))
        pb = cur.fetchone()
        batch_name = pb['batch_name'] if pb else batch_id

        if not rows:
            flash(f"Batch '{batch_name}' has no orders — nothing to undo.", 'warning')
            return redirect(request.referrer or url_for('admin.batch_history'))

        eligible_ids = [str(r['transaction_id']) for r in rows if r['is_editable']]
        locked       = [r for r in rows if not r['is_editable']]

        if eligible_ids:
            cur.execute("""
                UPDATE transactions
                SET print_batch_id = NULL, print_date = NULL,
                    fulfillment_status = 'uploaded', fulfillment_status_updated_at = NOW()
                WHERE transaction_id = ANY(%s::uuid[])
            """, (eligible_ids,))

            # Drop any receiving_items this batch's OPEN session was still holding
            # for the orders we just released, so the session doesn't keep
            # showing phantom rows for orders no longer in the batch.
            cur.execute("""
                DELETE FROM receiving_items ri
                USING receiving_sessions rs
                WHERE ri.session_id = rs.session_id
                  AND rs.status = 'open'
                  AND rs.batch_id = %s
                  AND ri.transaction_id = ANY(%s::uuid[])
            """, (batch_id, eligible_ids))

        # If nothing's left in the batch, the metadata row can go too.
        cur.execute("""
            DELETE FROM print_batches
            WHERE batch_id = %s
              AND NOT EXISTS (SELECT 1 FROM transactions WHERE print_batch_id = %s)
        """, (batch_id, batch_id))

    detail = f"{len(eligible_ids)} of {len(rows)} order(s) reverted to unprinted"
    if locked:
        detail += "; kept in batch: " + ", ".join(f"{r['order_number']} ({r['lock_reason']})" for r in locked)
    audit('batch_undone', 'print_batch', batch_id, detail=f"'{batch_name}' — {detail}")

    if locked:
        locked_desc = ", ".join(f"{r['order_number']} ({r['lock_reason']})" for r in locked)
        flash(f"Batch '{batch_name}': {len(eligible_ids)} of {len(rows)} order(s) released. "
              f"Left in place — {locked_desc}.", 'warning')
    else:
        flash(f"Batch '{batch_name}' undone. All {len(eligible_ids)} order(s) returned to the unprinted queue.",
              'success')
    return redirect(request.referrer or url_for('admin.batch_history'))


@admin_bp.route('/batch/<batch_id>/rename', methods=['POST'])
@login_required
@require_role('admin')
def batch_rename(batch_id):
    from ..security import audit
    from flask import abort

    new_name = request.form.get('batch_name', '').strip()
    if not new_name:
        flash('Batch name cannot be empty.', 'error')
        return redirect(request.referrer or url_for('admin.batch_history'))

    with db_cursor() as (cur, conn):
        cur.execute("SELECT batch_name FROM print_batches WHERE batch_id = %s", (batch_id,))
        pb = cur.fetchone()
        if not pb:
            abort(404)
        old_name = pb['batch_name']

        cur.execute("""
            SELECT 1 FROM print_batches
            WHERE LOWER(batch_name) = LOWER(%s) AND batch_id != %s
        """, (new_name, batch_id))
        if cur.fetchone():
            flash(f"Another batch is already named '{new_name}'. Pick something else.", 'error')
            return redirect(request.referrer or url_for('admin.batch_history'))

        cur.execute("""
            UPDATE print_batches SET batch_name = %s, updated_at = NOW() WHERE batch_id = %s
        """, (new_name, batch_id))

    audit('batch_renamed', 'print_batch', batch_id, detail=f"'{old_name}' → '{new_name}'")
    flash(f"Batch renamed to '{new_name}'.", 'success')
    return redirect(request.referrer or url_for('admin.batch_history'))


@admin_bp.route('/batch/<batch_id>/remove-order', methods=['POST'])
@login_required
@require_role('admin')
def batch_remove_order(batch_id):
    from ..security import audit

    tid = request.form.get('transaction_id', '').strip()
    if not tid:
        return redirect(url_for('admin.batch_detail', batch_id=batch_id))

    with db_cursor() as (cur, conn):
        cur.execute(f"""
            SELECT t.order_number, {_EDITABLE_EXPR} AS is_editable, {_LOCK_REASON_EXPR} AS lock_reason
            FROM transactions t WHERE t.transaction_id = %s AND t.print_batch_id = %s
        """, (tid, batch_id))
        row = cur.fetchone()
        if not row:
            flash('Order not found in this batch.', 'error')
            return redirect(url_for('admin.batch_detail', batch_id=batch_id))

        if not row['is_editable']:
            flash(f"Can't remove {row['order_number']} — {row['lock_reason']}.", 'warning')
            return redirect(url_for('admin.batch_detail', batch_id=batch_id))

        cur.execute("""
            UPDATE transactions
            SET print_batch_id = NULL, print_date = NULL,
                fulfillment_status = 'uploaded', fulfillment_status_updated_at = NOW()
            WHERE transaction_id = %s
        """, (tid,))

        cur.execute("""
            DELETE FROM receiving_items ri
            USING receiving_sessions rs
            WHERE ri.session_id = rs.session_id
              AND rs.status = 'open' AND rs.batch_id = %s
              AND ri.transaction_id = %s
        """, (batch_id, tid))

        cur.execute("""
            DELETE FROM print_batches
            WHERE batch_id = %s
              AND NOT EXISTS (SELECT 1 FROM transactions WHERE print_batch_id = %s)
        """, (batch_id, batch_id))

    audit('batch_order_removed', 'transaction', tid, detail=f"Removed {row['order_number']} from batch {batch_id}")
    flash(f"{row['order_number']} removed from batch and returned to the unprinted queue.", 'success')
    return redirect(url_for('admin.batch_detail', batch_id=batch_id))


@admin_bp.route('/batch/<batch_id>/add-order', methods=['POST'])
@login_required
@require_role('admin')
def batch_add_order(batch_id):
    from ..security import audit
    from flask_login import current_user
    from .receiving import _add_orphaned_batch_orders_to_session

    tid = request.form.get('transaction_id', '').strip()
    if not tid:
        return redirect(url_for('admin.batch_detail', batch_id=batch_id))

    with db_cursor() as (cur, conn):
        cur.execute("""
            UPDATE transactions
            SET print_batch_id = %s, print_date = NOW(),
                fulfillment_status = 'batched', fulfillment_status_updated_at = NOW()
            WHERE transaction_id = %s AND fulfillment_status = 'uploaded' AND is_active = TRUE
            RETURNING order_number
        """, (batch_id, tid))
        updated = cur.fetchone()

        if not updated:
            flash("That order is no longer available to add (it may have just been batched elsewhere).", 'error')
            return redirect(url_for('admin.batch_detail', batch_id=batch_id))

        # Make sure this batch has a metadata row (covers the edge case of
        # adding to a batch_id that was fully undone down to zero orders).
        safe_name = _unique_batch_name(cur, batch_id, batch_id)
        cur.execute("""
            INSERT INTO print_batches (batch_id, batch_name, created_by)
            VALUES (%s, %s, %s)
            ON CONFLICT (batch_id) DO NOTHING
        """, (batch_id, safe_name, current_user.id if current_user.is_authenticated else None))

        # If a receiving session (open or closed) already exists for this batch,
        # sync the newly-added order into it — same helper used elsewhere for
        # exactly this "order joined the batch after the session started" case.
        cur.execute("SELECT session_id FROM receiving_sessions WHERE batch_id = %s", (batch_id,))
        for sess in cur.fetchall():
            _add_orphaned_batch_orders_to_session(cur, sess['session_id'], batch_id)

    audit('batch_order_added', 'transaction', tid, detail=f"Added {updated['order_number']} to batch {batch_id}")
    flash(f"{updated['order_number']} added to the batch.", 'success')
    return redirect(url_for('admin.batch_detail', batch_id=batch_id))


@admin_bp.route('/print-invoice/<string:tid>')
@login_required
@require_role('admin')
def print_invoice(tid):
    from flask import send_file, abort, redirect as redir, make_response
    import io, os
    from datetime import datetime

    # Optional batch context passed as query params for watermark
    batch_id     = request.args.get('batch_id', '')
    company_name = request.args.get('company', '')

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT t.invoice_file_path, t.invoice_pdf, t.print_batch_id,
                   c.company_name, t.order_number
            FROM transactions t
            LEFT JOIN dim_companies c ON t.company_id = c.company_id
            WHERE t.transaction_id = %s
        """, (str(tid),))
        row = cur.fetchone()
    if not row:
        abort(404)

    # Use batch_id from query param or from DB
    batch  = batch_id or row['print_batch_id'] or ''
    comp   = company_name or row['company_name'] or ''
    today  = datetime.now().strftime('%b %d, %Y')

    # Priority 1: PDF stored in DB — stamp watermark if batch context
    if row['invoice_pdf']:
        pdf_bytes = bytes(row['invoice_pdf'])
        if batch:
            try:
                from ..watermark import stamp_pdf
                pdf_bytes = stamp_pdf(pdf_bytes, batch_id=batch,
                                      company_name=comp, print_date=today)
            except Exception:
                pass  # serve unstamped if watermark fails
        fname = f"invoice-{row['order_number'] or tid[:8]}.pdf"
        return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                         download_name=fname)

    # Priority 2: Google Drive / HTTP link (can't stamp, just redirect)
    path = row['invoice_file_path'] or ''
    if path.startswith('http'):
        return redir(path)

    # Priority 3: Local file
    if path and os.path.exists(path):
        return send_file(path, mimetype='application/pdf')

    return make_response(
        "<h2>Invoice PDF unavailable</h2><p>This invoice was submitted before PDF "
        "storage was enabled.</p><a href='javascript:history.back()'>← Go back</a>", 404)


@admin_bp.route('/payroll/export')
@login_required
@require_role('admin')
def export_payroll():
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl not installed. Add it to requirements.txt.', 'error')
        return redirect(url_for('admin.payroll'))
    from flask import send_file as sf

    month   = request.args.get('month', '')
    company = request.args.get('company', type=int)

    conditions = ["t.review_status != 'Flagged'", "t.is_duplicate=FALSE",
                  "t.price_total>0", "t.is_active=TRUE"]
    params = []
    if month:
        conditions.append("t.purchase_year_month=%s"); params.append(month)
    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1a1d27')

    for comp in companies:
        if company and comp['company_id'] != company:
            continue
        with db_cursor() as (cur, _):
            cur.execute(f"""
                SELECT u.username, t.purchase_year_month,
                       COUNT(*) AS orders,
                       ROUND(SUM(t.price_total)::numeric,2) AS purchases,
                       ROUND(SUM(COALESCE(t.gross_paid_amount,0))::numeric,2) AS gross_paid,
                       ROUND(SUM(COALESCE(t.net_paid_amount,0))::numeric,2) AS net_paid,
                       ROUND(SUM(COALESCE(t.sales_payroll_tax_withheld,0))::numeric,2) AS tax_withheld
                FROM transactions t
                LEFT JOIN dim_users u ON t.user_id=u.user_id
                {where} AND t.company_id=%s
                GROUP BY u.username, t.purchase_year_month
                ORDER BY t.purchase_year_month DESC, u.username
            """, params + [comp['company_id']])
            rows = cur.fetchall()

        if not rows:
            continue

        ws = wb.create_sheet(title=comp['company_name'][:31])
        headers = ['Person', 'Month', 'Orders', 'Total Purchases', 'Gross Paid (1%)', 'Net Paid (0.8%)', 'Tax Withheld (0.2%)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=r['username'] or '—')
            ws.cell(row=row_idx, column=2, value=r['purchase_year_month'])
            ws.cell(row=row_idx, column=3, value=r['orders'])
            for col, key in enumerate(['purchases','gross_paid','net_paid','tax_withheld'], 4):
                cell = ws.cell(row=row_idx, column=col, value=float(r[key] or 0))
                cell.number_format = '$#,##0.00'

        # Totals row
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=f'=SUM(C2:C{total_row-1})').font = Font(bold=True)
        for col in range(4, 8):
            col_letter = chr(64+col)
            ws.cell(row=total_row, column=col, value=f'=SUM({col_letter}2:{col_letter}{total_row-1})').font = Font(bold=True)
            ws.cell(row=total_row, column=col).number_format = '$#,##0.00'

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    fname = f"payroll{'_'+month if month else ''}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return sf(buf, as_attachment=True, download_name=fname,
              mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin_bp.route('/storage')
@login_required
@require_role('admin')
def storage():
    with db_cursor() as (cur, _):
        # Total DB size
        cur.execute("SELECT pg_database_size(current_database()) AS bytes")
        db_bytes = cur.fetchone()['bytes']

        # Per-table sizes
        cur.execute("""
            SELECT tablename,
                   pg_total_relation_size(schemaname||'.'||tablename) AS bytes,
                   pg_size_pretty(pg_total_relation_size(schemaname||'.'||tablename)) AS size
            FROM pg_tables WHERE schemaname='public'
            ORDER BY bytes DESC
        """)
        tables = cur.fetchall()

        # PDF storage specifically
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE invoice_pdf IS NOT NULL) AS pdf_count,
                   COALESCE(SUM(LENGTH(invoice_pdf)),0) AS pdf_bytes
            FROM transactions
        """)
        pdf_stats = cur.fetchone()

        # Transaction count over time
        cur.execute("""
            SELECT purchase_year_month,
                   COUNT(*) AS orders,
                   ROUND(SUM(price_total)::numeric,2) AS gmv
            FROM transactions
            WHERE is_active=TRUE AND price_total > 0
            GROUP BY purchase_year_month
            ORDER BY purchase_year_month DESC
            LIMIT 12
        """)
        monthly = cur.fetchall()

    # Railway Hobby plan = 5GB
    db_limit = 5368709120  # 5GB

    return render_template('storage.html',
                           db_bytes=db_bytes, db_limit=db_limit,
                           tables=tables, pdf_stats=pdf_stats, monthly=monthly)


@admin_bp.route('/transaction-pdf/<string:tid>')
@login_required
@require_role('admin')
def serve_transaction_pdf(tid):
    from flask import send_file, abort
    import io
    with db_cursor() as (cur, _):
        cur.execute("SELECT invoice_pdf FROM transactions WHERE transaction_id=%s", (str(tid),))
        row = cur.fetchone()
    if not row or not row['invoice_pdf']:
        abort(404)
    return send_file(io.BytesIO(bytes(row['invoice_pdf'])),
                     mimetype='application/pdf',
                     download_name=f'invoice-{tid[:8]}.pdf')

STALE_DAYS = 30

@admin_bp.route('/costco-taxes')
@login_required
@require_role('admin')
def costco_taxes():
    f_month      = request.args.get('month', '')
    f_year       = request.args.get('year', '')
    f_company    = request.args.get('company', type=int)
    f_person     = request.args.get('person_by', type=int)
    f_membership = request.args.get('membership', '')
    f_status     = request.args.get('status', '')
    f_order      = request.args.get('order_number', '').strip()

    conditions = ["t.retailer = 'Costco'", "t.is_active = TRUE", "t.exception_status IS DISTINCT FROM 'returned'"]
    params = []
    if f_month:
        conditions.append("TO_CHAR(t.purchase_date,'MM') = %s"); params.append(f_month)
    if f_year:
        conditions.append("TO_CHAR(t.purchase_date,'YYYY') = %s"); params.append(f_year)
    if f_company:
        conditions.append("t.company_id = %s"); params.append(f_company)
    if f_person:
        conditions.append("t.user_id = %s"); params.append(f_person)
    if f_membership:
        conditions.append("t.membership_number = %s"); params.append(f_membership)
    if f_order:
        conditions.append("t.order_number ILIKE %s"); params.append(f'%{f_order}%')
    if f_status == 'unrequested':
        conditions.append("t.costco_refund_status IS NULL")
    elif f_status == 'stale':
        conditions.append("t.costco_refund_status IN ('Pending','Partial')")
        conditions.append("t.costco_last_activity_at < NOW() - INTERVAL '%s days'" % STALE_DAYS)
    elif f_status in ('Pending', 'Partial', 'Full'):
        conditions.append("t.costco_refund_status = %s"); params.append(f_status)

    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        # KPIs
        cur.execute(f"""
            SELECT
                COUNT(*)                                                        AS total_orders,
                ROUND(SUM(t.price_total)::numeric, 2)                          AS total_gmv,
                ROUND(SUM(COALESCE(t.costco_taxes_paid,0))::numeric, 2)        AS total_taxes,
                ROUND(SUM(COALESCE(t.gross_paid_amount,0))::numeric, 2)        AS total_gross_paid,
                ROUND(SUM(COALESCE(t.cashback_value,0))::numeric, 2)           AS total_cashback,
                COUNT(DISTINCT t.membership_number) FILTER
                    (WHERE t.membership_number IS NOT NULL)                     AS unique_memberships,
                COUNT(*) FILTER (WHERE t.costco_refund_status IS NULL)          AS unrequested_count,
                COUNT(*) FILTER (WHERE t.costco_refund_status = 'Pending')      AS pending_count,
                COUNT(*) FILTER (WHERE t.costco_refund_status = 'Partial')      AS partial_count,
                COUNT(*) FILTER (WHERE t.costco_refund_status = 'Full')         AS full_count,
                COUNT(*) FILTER (WHERE t.costco_refund_status IN ('Pending','Partial')
                                   AND t.costco_last_activity_at < NOW() - INTERVAL '{STALE_DAYS} days') AS stale_count,
                ROUND((SUM(COALESCE(t.costco_taxes_paid,0) - COALESCE(t.costco_refund_amount,0))
                       FILTER (WHERE t.costco_refund_status IS DISTINCT FROM 'Full'))::numeric, 2) AS outstanding_tax
            FROM transactions t {where}
        """, params)
        metrics = cur.fetchone()

        # Transactions table
        cur.execute(f"""
            SELECT t.transaction_id, t.order_number, t.purchase_date,
                   t.price_total, t.costco_taxes_paid,
                   t.membership_number, t.card_id, d.cashback_rate,
                   ROUND(COALESCE(t.gross_paid_amount,0)::numeric,2)     AS gross_paid,
                   ROUND(COALESCE(t.net_paid_amount,0)::numeric,2)       AS net_paid,
                   ROUND(COALESCE(t.sales_payroll_tax_withheld,0)::numeric,2) AS tax_withheld,
                   ROUND(COALESCE(t.cashback_value,0)::numeric,2)        AS cashback,
                   (t.invoice_pdf IS NOT NULL)                           AS has_pdf,
                   t.invoice_file_path,
                   per.username AS person_name, c.company_name, t.company_id,
                   t.review_status,
                   t.costco_refund_status, t.costco_refund_amount,
                   t.costco_last_requested_at, t.costco_last_activity_at,
                   ROUND((COALESCE(t.costco_taxes_paid,0) - COALESCE(t.costco_refund_amount,0))::numeric, 2) AS missing_amount,
                   CASE WHEN t.costco_last_activity_at IS NOT NULL
                        THEN EXTRACT(DAY FROM NOW() - t.costco_last_activity_at)::int
                        ELSE NULL END AS days_waiting,
                   (SELECT b.batch_name FROM costco_tax_batch_items bi
                      JOIN costco_tax_batches b ON bi.batch_id = b.batch_id
                     WHERE bi.transaction_id = t.transaction_id
                     ORDER BY b.created_at DESC LIMIT 1)                 AS last_batch_name,
                   (SELECT b.batch_id FROM costco_tax_batch_items bi
                      JOIN costco_tax_batches b ON bi.batch_id = b.batch_id
                     WHERE bi.transaction_id = t.transaction_id
                     ORDER BY b.created_at DESC LIMIT 1)                 AS last_batch_id
            FROM transactions t
            LEFT JOIN dim_users per    ON t.user_id    = per.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            LEFT JOIN dim_cards d      ON t.card_id    = d.card_id
            {where}
            ORDER BY t.purchase_date DESC
            LIMIT 200
        """, params)
        transactions = cur.fetchall()

        # Filter options
        cur.execute("SELECT DISTINCT TO_CHAR(purchase_date,'YYYY') AS yr FROM transactions WHERE purchase_date IS NOT NULL ORDER BY yr DESC")
        years = [r['yr'] for r in cur.fetchall()]
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT user_id, username FROM dim_users WHERE is_active=TRUE ORDER BY username")
        users = cur.fetchall()
        cur.execute("SELECT DISTINCT membership_number FROM transactions WHERE membership_number IS NOT NULL ORDER BY membership_number")
        memberships = [r['membership_number'] for r in cur.fetchall()]

    return render_template('costco_taxes.html',
                           metrics=metrics, transactions=transactions,
                           years=years, companies=companies, users=users, memberships=memberships,
                           stale_days=STALE_DAYS,
                           filters={'month':f_month,'year':f_year,'company':f_company,
                                    'person_by':f_person,'membership':f_membership,
                                    'status':f_status,'order_number':f_order})


@admin_bp.route('/costco-taxes/batch/create', methods=['POST'])
@login_required
@require_role('admin')
def costco_tax_batch_create():
    from ..security import audit
    from flask_login import current_user
    import uuid as _uuid

    txn_ids    = request.form.getlist('txn_ids')
    batch_name = request.form.get('batch_name', '').strip()
    company_id = request.form.get('company_id', type=int)

    if not txn_ids or not batch_name:
        flash('Pick at least one order and give the batch a name.', 'danger')
        return redirect(url_for('admin.costco_taxes'))

    with db_cursor() as (cur, conn):
        cur.execute("""
            SELECT transaction_id, order_number, costco_taxes_paid, costco_refund_amount,
                   costco_refund_status
            FROM transactions
            WHERE transaction_id = ANY(%s::uuid[]) AND retailer = 'Costco'
        """, (txn_ids,))
        rows = cur.fetchall()

        eligible = [r for r in rows if r['costco_refund_status'] != 'Full']
        if not eligible:
            flash('None of the selected orders are eligible (already fully refunded).', 'danger')
            return redirect(url_for('admin.costco_taxes'))

        batch_id = str(_uuid.uuid4())
        total_requested = 0.0
        item_rows = []
        for r in eligible:
            tax_paid = float(r['costco_taxes_paid'] or 0)
            refunded = float(r['costco_refund_amount'] or 0)
            amount_requested = round(tax_paid - refunded, 2)
            if amount_requested <= 0:
                continue
            total_requested += amount_requested
            item_rows.append((batch_id, str(r['transaction_id']), amount_requested))

        if not item_rows:
            flash('Selected orders have nothing outstanding to request.', 'danger')
            return redirect(url_for('admin.costco_taxes'))

        cur.execute("""
            INSERT INTO costco_tax_batches
                (batch_id, batch_name, company_id, created_by, order_count, total_requested)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (batch_id, batch_name, company_id, current_user.id if current_user.is_authenticated else None,
              len(item_rows), round(total_requested, 2)))

        cur.executemany("""
            INSERT INTO costco_tax_batch_items (batch_id, transaction_id, amount_requested)
            VALUES (%s, %s, %s)
        """, item_rows)

        included_ids = [r[1] for r in item_rows]
        cur.execute("""
            UPDATE transactions
            SET costco_refund_status = COALESCE(costco_refund_status, 'Pending'),
                costco_last_requested_at = NOW(),
                costco_last_activity_at = NOW()
            WHERE transaction_id = ANY(%s::uuid[])
        """, (included_ids,))

    audit('costco_batch_created', 'costco_tax_batch', batch_id,
          detail=f"'{batch_name}' — {len(item_rows)} orders, ${total_requested:,.2f} requested")
    flash(f"Batch '{batch_name}' created — {len(item_rows)} orders, ${total_requested:,.2f} requested.", 'success')
    return redirect(url_for('admin.costco_tax_batch_download', batch_id=batch_id))


@admin_bp.route('/costco-taxes/batch/<batch_id>/download')
@login_required
@require_role('admin')
def costco_tax_batch_download(batch_id):
    import io
    from flask import send_file as sf, abort
    import openpyxl
    from openpyxl.styles import Font

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT b.batch_name, b.created_at, c.company_name
            FROM costco_tax_batches b
            LEFT JOIN dim_companies c ON b.company_id = c.company_id
            WHERE b.batch_id = %s
        """, (batch_id,))
        batch = cur.fetchone()
        if not batch:
            abort(404)

        cur.execute("""
            SELECT t.membership_number, per.username AS person_name, t.purchase_date,
                   t.order_number, t.costco_ship_city, t.costco_ship_state,
                   bi.amount_requested
            FROM costco_tax_batch_items bi
            JOIN transactions t ON bi.transaction_id = t.transaction_id
            LEFT JOIN dim_users per ON t.user_id = per.user_id
            WHERE bi.batch_id = %s
            ORDER BY t.purchase_date
        """, (batch_id,))
        rows = cur.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Costco Tax Request'

    title = f"{batch['company_name'] or ''} / {batch['batch_name']}".strip(' /')
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    headers = ['Membership', 'Name', 'Date', 'Order', 'Ship to City', 'Ship to State', 'Pending Tax']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)

    total = 0.0
    for ri, r in enumerate(rows, 3):
        amt = float(r['amount_requested'] or 0)
        total += amt
        vals = [r['membership_number'] or '', r['person_name'] or '',
                r['purchase_date'], r['order_number'],
                r['costco_ship_city'] or 'Doral', r['costco_ship_state'] or 'FL', amt]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci == 7:
                cell.number_format = '$#,##0.00'

    tr = len(rows) + 3
    ws.cell(row=tr, column=1, value='Total to be refunded').font = Font(bold=True)
    ws.cell(row=tr, column=7, value=round(total, 2)).font = Font(bold=True)
    ws.cell(row=tr, column=7).number_format = '$#,##0.00'

    widths = [16, 20, 12, 14, 16, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    fname = f"costco_tax_request_{batch['batch_name']}.xlsx".replace(' ', '_')
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return sf(buf, as_attachment=True, download_name=fname,
              mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/costco-taxes/batches')
@login_required
@require_role('admin')
def costco_tax_batches():
    f_company = request.args.get('company', type=int)
    f_batch   = request.args.get('batch_name', '').strip()

    conditions, params = [], []
    if f_company:
        conditions.append("b.company_id = %s"); params.append(f_company)
    if f_batch:
        conditions.append("b.batch_name ILIKE %s"); params.append(f'%{f_batch}%')
    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT b.batch_id, b.batch_name, b.created_at, b.order_count, b.total_requested,
                   c.company_name, u.username AS created_by,
                   COUNT(*) FILTER (WHERE t.costco_refund_status = 'Full')    AS full_count,
                   COUNT(*) FILTER (WHERE t.costco_refund_status = 'Partial') AS partial_count,
                   COUNT(*) FILTER (WHERE t.costco_refund_status = 'Pending') AS pending_count
            FROM costco_tax_batches b
            LEFT JOIN dim_companies c ON b.company_id = c.company_id
            LEFT JOIN dim_users u     ON b.created_by  = u.user_id
            LEFT JOIN costco_tax_batch_items bi ON bi.batch_id = b.batch_id
            LEFT JOIN transactions t ON bi.transaction_id = t.transaction_id
            {where}
            GROUP BY b.batch_id, b.batch_name, b.created_at, b.order_count, b.total_requested,
                     c.company_name, u.username
            ORDER BY b.created_at DESC
        """, params)
        batches = cur.fetchall()
        cur.execute("SELECT company_id, company_name FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()

    return render_template('costco_tax_batches.html', batches=batches, companies=companies,
                           filters={'company': f_company, 'batch_name': f_batch})


@admin_bp.route('/costco-taxes/batch/<batch_id>')
@login_required
@require_role('admin')
def costco_tax_batch_detail(batch_id):
    from flask import abort
    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT b.batch_id, b.batch_name, b.created_at, b.total_requested,
                   c.company_name, u.username AS created_by
            FROM costco_tax_batches b
            LEFT JOIN dim_companies c ON b.company_id = c.company_id
            LEFT JOIN dim_users u     ON b.created_by  = u.user_id
            WHERE b.batch_id = %s
        """, (batch_id,))
        batch = cur.fetchone()
        if not batch:
            abort(404)

        cur.execute("""
            SELECT t.transaction_id, t.order_number, t.purchase_date, per.username AS person_name,
                   t.costco_taxes_paid, t.costco_refund_status, t.costco_refund_amount,
                   bi.amount_requested,
                   ROUND((COALESCE(t.costco_taxes_paid,0) - COALESCE(t.costco_refund_amount,0))::numeric,2) AS missing_amount,
                   CASE WHEN t.costco_last_activity_at IS NOT NULL
                        THEN EXTRACT(DAY FROM NOW() - t.costco_last_activity_at)::int
                        ELSE NULL END AS days_waiting
            FROM costco_tax_batch_items bi
            JOIN transactions t ON bi.transaction_id = t.transaction_id
            LEFT JOIN dim_users per ON t.user_id = per.user_id
            WHERE bi.batch_id = %s
            ORDER BY t.purchase_date
        """, (batch_id,))
        orders = cur.fetchall()

    return render_template('costco_tax_batch_detail.html', batch=batch, orders=orders)


@admin_bp.route('/costco-taxes/refund-update', methods=['POST'])
@login_required
@require_role('admin')
def costco_tax_refund_update():
    from ..security import audit
    from flask import abort

    tid           = request.form.get('transaction_id', '').strip()
    status_choice = request.form.get('status_choice', 'Pending')
    amount_entry  = request.form.get('amount_received', type=float)
    return_to     = request.form.get('return_to') or url_for('admin.costco_taxes')

    if status_choice not in ('Pending', 'Partial', 'Full'):
        status_choice = 'Pending'

    if not tid:
        flash('Missing order reference.', 'danger')
        return redirect(return_to)

    with db_cursor() as (cur, conn):
        cur.execute("""
            SELECT order_number, costco_taxes_paid
            FROM transactions WHERE transaction_id = %s AND retailer = 'Costco'
        """, (tid,))
        row = cur.fetchone()
        if not row:
            abort(404)

        tax_paid = float(row['costco_taxes_paid'] or 0)

        if status_choice == 'Full':
            new_amount = tax_paid
            new_status = 'Full'
        elif status_choice == 'Partial':
            if amount_entry is None or amount_entry <= 0:
                flash('Enter a partial refund amount greater than $0.', 'danger')
                return redirect(return_to)
            new_amount = round(min(amount_entry, tax_paid), 2)
            new_status = 'Partial'
        else:  # Pending — nothing selected, or explicitly reset
            new_amount = 0
            new_status = 'Pending'

        cur.execute("""
            UPDATE transactions
            SET costco_refund_amount = %s,
                costco_refund_status = %s,
                costco_last_activity_at = NOW()
            WHERE transaction_id = %s
        """, (new_amount, new_status, tid))

    audit('costco_refund_updated', 'transaction', tid,
          detail=f"Order {row['order_number']}: marked {new_status}"
                 + (f" — ${new_amount:,.2f} of ${tax_paid:,.2f}" if new_status != 'Pending' else ""))
    flash(f"Order {row['order_number']} marked {new_status}.", 'success')
    return redirect(return_to)


@admin_bp.route('/costco-taxes/batch/<batch_id>/undo', methods=['POST'])
@login_required
@require_role('admin')
def costco_tax_batch_undo(batch_id):
    from ..security import audit
    from flask import abort

    with db_cursor() as (cur, conn):
        cur.execute("SELECT batch_name FROM costco_tax_batches WHERE batch_id = %s", (batch_id,))
        batch = cur.fetchone()
        if not batch:
            abort(404)

        # Orders eligible to revert to Unrequested: this batch was their only link.
        # Reverts regardless of current status (Pending/Partial/Full) — undoing a
        # batch/order-removal resets it fully rather than preserving a refund that
        # was only ever recorded in the context of this batch.
        cur.execute("""
            SELECT bi.transaction_id, t.costco_refund_status,
                   (SELECT COUNT(*) FROM costco_tax_batch_items bi2
                     WHERE bi2.transaction_id = bi.transaction_id AND bi2.batch_id != %s) AS other_batches
            FROM costco_tax_batch_items bi
            JOIN transactions t ON bi.transaction_id = t.transaction_id
            WHERE bi.batch_id = %s
        """, (batch_id, batch_id))
        items = cur.fetchall()

        revert_ids = [str(i['transaction_id']) for i in items
                      if i['other_batches'] == 0]

        if revert_ids:
            cur.execute("""
                UPDATE transactions
                SET costco_refund_status = NULL,
                    costco_refund_amount = 0,
                    costco_last_requested_at = NULL,
                    costco_last_activity_at = NULL
                WHERE transaction_id = ANY(%s::uuid[])
            """, (revert_ids,))

        cur.execute("DELETE FROM costco_tax_batches WHERE batch_id = %s", (batch_id,))
        # costco_tax_batch_items rows cascade-delete via FK

    audit('costco_batch_undone', 'costco_tax_batch', batch_id,
          detail=f"Undid batch '{batch['batch_name']}' — {len(revert_ids)} order(s) reverted to Unrequested")
    flash(f"Batch '{batch['batch_name']}' undone. {len(revert_ids)} order(s) reverted to Unrequested; "
          f"any order still linked to another batch was left untouched.", 'success')
    return redirect(url_for('admin.costco_tax_batches'))


@admin_bp.route('/costco-taxes/batch/<batch_id>/remove-order', methods=['POST'])
@login_required
@require_role('admin')
def costco_tax_batch_remove_order(batch_id):
    from ..security import audit

    tid = request.form.get('transaction_id', '').strip()
    if not tid:
        flash('Missing order reference.', 'danger')
        return redirect(url_for('admin.costco_tax_batch_detail', batch_id=batch_id))

    with db_cursor() as (cur, conn):
        cur.execute("""
            SELECT COUNT(*) AS n FROM costco_tax_batch_items
            WHERE transaction_id = %s AND batch_id != %s
        """, (tid, batch_id))
        other_batches = cur.fetchone()['n']

        cur.execute("SELECT order_number, costco_refund_status FROM transactions WHERE transaction_id = %s", (tid,))
        t = cur.fetchone()

        cur.execute("DELETE FROM costco_tax_batch_items WHERE batch_id = %s AND transaction_id = %s",
                    (batch_id, tid))

        cur.execute("""
            SELECT COUNT(*) AS cnt, COALESCE(SUM(amount_requested), 0) AS total
            FROM costco_tax_batch_items WHERE batch_id = %s
        """, (batch_id,))
        agg = cur.fetchone()
        cur.execute("""
            UPDATE costco_tax_batches SET order_count = %s, total_requested = %s WHERE batch_id = %s
        """, (agg['cnt'], round(float(agg['total'] or 0), 2), batch_id))

        reverted = False
        if other_batches == 0:
            cur.execute("""
                UPDATE transactions
                SET costco_refund_status = NULL, costco_refund_amount = 0,
                    costco_last_requested_at = NULL, costco_last_activity_at = NULL
                WHERE transaction_id = %s
            """, (tid,))
            reverted = True

    audit('costco_batch_order_removed', 'transaction', tid,
          detail=f"Removed from batch {batch_id}" + (" — reverted to Unrequested" if reverted else ""))
    flash(f"Order {t['order_number'] if t else tid} removed from batch.", 'success')
    return redirect(url_for('admin.costco_tax_batch_detail', batch_id=batch_id))


@admin_bp.route('/costco-taxes/bulk-complete', methods=['POST'])
@login_required
@require_role('admin')
def costco_tax_bulk_complete():
    from ..security import audit

    txn_ids   = request.form.getlist('txn_ids')
    return_to = request.form.get('return_to') or url_for('admin.costco_taxes')

    if not txn_ids:
        flash('Select at least one order to mark complete.', 'danger')
        return redirect(return_to)

    with db_cursor() as (cur, conn):
        cur.execute("""
            UPDATE transactions
            SET costco_refund_amount = COALESCE(costco_taxes_paid, 0),
                costco_refund_status = 'Full',
                costco_last_activity_at = NOW()
            WHERE transaction_id = ANY(%s::uuid[]) AND retailer = 'Costco'
            RETURNING transaction_id, order_number
        """, (txn_ids,))
        updated = cur.fetchall()

    for u in updated:
        audit('costco_bulk_marked_complete', 'transaction', str(u['transaction_id']),
              detail=f"Order {u['order_number']}: bulk-marked Complete")
    flash(f"{len(updated)} order(s) marked Complete.", 'success')
    return redirect(return_to)


@admin_bp.route('/audit-log')
@login_required
@require_role('admin')
def audit_log():
    page     = request.args.get('page', 1, type=int)
    per_page = 50
    offset   = (page - 1) * per_page
    f_action = request.args.get('action', '')
    f_user   = request.args.get('user', '')

    conditions, params = [], []
    if f_action:
        conditions.append("a.action = %s"); params.append(f_action)
    if f_user:
        conditions.append("a.user_email ILIKE %s"); params.append(f'%{f_user}%')

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT a.log_id, a.action, a.target_type, a.target_id,
                   a.detail, a.ip_address, a.created_at,
                   a.user_email, u.username
            FROM audit_log a
            LEFT JOIN dim_users u ON a.user_id = u.user_id
            {where}
            ORDER BY a.created_at DESC LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        logs = cur.fetchall()
        cur.execute(f"SELECT COUNT(*) AS n FROM audit_log a {where}", params)
        total = cur.fetchone()['n']
        cur.execute("SELECT DISTINCT action FROM audit_log ORDER BY action")
        actions = [r['action'] for r in cur.fetchall()]

    return render_template('audit_log.html', logs=logs, total=total,
                           page=page, per_page=per_page, actions=actions,
                           filters={'action': f_action, 'user': f_user})

@admin_bp.route('/costco-taxes/export')
@login_required
@require_role('admin')
def export_costco_taxes():
    import io
    from flask import send_file as sf
    fmt = request.args.get('fmt', 'excel')

    f_month      = request.args.get('month', '')
    f_year       = request.args.get('year', '')
    f_company    = request.args.get('company', type=int)
    f_person     = request.args.get('person_by', type=int)
    f_membership = request.args.get('membership', '')

    conditions = ["t.retailer = 'Costco'", "t.is_active = TRUE", "t.exception_status IS DISTINCT FROM 'returned'"]
    params = []
    if f_month:
        conditions.append("TO_CHAR(t.purchase_date,'MM') = %s"); params.append(f_month)
    if f_year:
        conditions.append("TO_CHAR(t.purchase_date,'YYYY') = %s"); params.append(f_year)
    if f_company:
        conditions.append("t.company_id = %s"); params.append(f_company)
    if f_person:
        conditions.append("t.user_id = %s"); params.append(f_person)
    if f_membership:
        conditions.append("t.membership_number = %s"); params.append(f_membership)
    where = 'WHERE ' + ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT t.order_number, t.purchase_date, per.username AS person_name,
                   c.company_name, t.membership_number, t.card_id,
                   ROUND(t.price_total::numeric,2)                              AS total,
                   ROUND(COALESCE(t.costco_taxes_paid,0)::numeric,2)            AS costco_tax,
                   ROUND(COALESCE(t.gross_paid_amount,0)::numeric,2)            AS gross_paid,
                   ROUND(COALESCE(t.net_paid_amount,0)::numeric,2)              AS net_paid,
                   ROUND(COALESCE(t.sales_payroll_tax_withheld,0)::numeric,2)   AS tax_withheld,
                   ROUND(COALESCE(t.cashback_value,0)::numeric,2)               AS cashback,
                   t.review_status
            FROM transactions t
            LEFT JOIN dim_users per    ON t.user_id    = per.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            {where}
            ORDER BY t.purchase_date DESC
        """, params)
        rows = cur.fetchall()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Costco Taxes'

    headers = ['Order #', 'Date', 'Person By', 'Company', 'Membership #',
               'Card', 'Total', 'Costco Tax', 'Gross Paid', 'Net Paid',
               'Tax Withheld', 'Cash Back', 'Status']
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', start_color='1a1d27')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    money_cols = {7,8,9,10,11,12}
    for ri, r in enumerate(rows, 2):
        vals = [r['order_number'], str(r['purchase_date']) if r['purchase_date'] else '',
                r['person_name'] or '', r['company_name'] or '',
                r['membership_number'] or '', r['card_id'] or '',
                float(r['total'] or 0), float(r['costco_tax'] or 0),
                float(r['gross_paid'] or 0), float(r['net_paid'] or 0),
                float(r['tax_withheld'] or 0), float(r['cashback'] or 0),
                r['review_status'] or '']
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if ci in money_cols:
                cell.number_format = '$#,##0.00'

    # Totals row
    tr = len(rows) + 2
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True)
    for ci, col in zip(range(7,13), 'GHIJKL'):
        ws.cell(row=tr, column=ci,
                value=f'=SUM({col}2:{col}{tr-1})').font = Font(bold=True)
        ws.cell(row=tr, column=ci).number_format = '$#,##0.00'

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    fname = f"costco_taxes{'_'+f_year if f_year else ''}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return sf(buf, as_attachment=True, download_name=fname,
              mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@admin_bp.route('/audit-log/export')
@login_required
@require_role('admin')
def export_audit_log():
    import io
    from flask import send_file as sf
    f_action = request.args.get('action', '')
    f_user   = request.args.get('user', '')

    conditions, params = [], []
    if f_action:
        conditions.append("a.action = %s"); params.append(f_action)
    if f_user:
        conditions.append("a.user_email ILIKE %s"); params.append(f'%{f_user}%')
    if f_batch:
        conditions.append("t.print_batch_id = %s"); params.append(f_batch)

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT a.created_at, a.action, u.username, a.user_email,
                   a.target_type, a.target_id, a.detail, a.ip_address
            FROM audit_log a
            LEFT JOIN dim_users u ON a.user_id = u.user_id
            {where}
            ORDER BY a.created_at DESC LIMIT 5000
        """, params)
        rows = cur.fetchall()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audit Log'
    headers = ['Timestamp', 'Action', 'Username', 'Email', 'Target Type', 'Target ID', 'Detail', 'IP Address']
    hfont = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', start_color='1a1d27')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=str(r['created_at']) if r['created_at'] else '')
        ws.cell(row=ri, column=2, value=r['action'] or '')
        ws.cell(row=ri, column=3, value=r['username'] or '')
        ws.cell(row=ri, column=4, value=r['user_email'] or '')
        ws.cell(row=ri, column=5, value=r['target_type'] or '')
        ws.cell(row=ri, column=6, value=r['target_id'] or '')
        ws.cell(row=ri, column=7, value=r['detail'] or '')
        ws.cell(row=ri, column=8, value=r['ip_address'] or '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return sf(buf, as_attachment=True, download_name='audit_log.xlsx',
              mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/batch/<batch_id>/print-all', methods=['POST'])
@login_required
@require_role('admin')
def batch_print_all(batch_id):
    """Merge all selected PDFs in a batch into one combined PDF with watermarks."""
    from flask import send_file, abort
    import io
    from datetime import datetime
    from pypdf import PdfReader, PdfWriter

    selected_tids = request.form.getlist('tids')
    if not selected_tids:
        abort(400)

    today = datetime.now().strftime('%b %d, %Y')

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT t.transaction_id, t.order_number, t.invoice_pdf,
                   t.invoice_file_path, c.company_name
            FROM transactions t
            LEFT JOIN dim_companies c ON t.company_id = c.company_id
            WHERE t.transaction_id = ANY(%s::uuid[])
              AND t.print_batch_id = %s
            ORDER BY t.purchase_date
        """, (selected_tids, batch_id))
        rows = cur.fetchall()

    if not rows:
        abort(404)

    writer = PdfWriter()
    included = 0

    for row in rows:
        pdf_bytes = None

        if row['invoice_pdf']:
            pdf_bytes = bytes(row['invoice_pdf'])
        elif row['invoice_file_path'] and not row['invoice_file_path'].startswith('http'):
            import os
            if os.path.exists(row['invoice_file_path']):
                with open(row['invoice_file_path'], 'rb') as f:
                    pdf_bytes = f.read()

        if not pdf_bytes:
            continue  # skip Drive links and missing PDFs

        # Stamp watermark on first page of each invoice
        try:
            from ..watermark import stamp_pdf
            company = row['company_name'] or ''
            pdf_bytes = stamp_pdf(pdf_bytes, batch_id=batch_id,
                                  company_name=company, print_date=today)
        except Exception:
            pass  # use unstamped if watermark fails

        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            for page in reader.pages:
                writer.add_page(page)
            included += 1
        except Exception:
            continue

    if included == 0:
        from flask import make_response
        return make_response(
            "<h2>No printable PDFs</h2><p>None of the selected invoices have PDFs stored in the database. "
            "Historical invoices with Google Drive links cannot be merged — open them individually with 🔗 Open.</p>"
            "<a href='javascript:history.back()'>← Go back</a>", 404)

    out = io.BytesIO()
    writer.write(out)
    out.seek(0)

    fname = f"batch-{batch_id}-{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(out, mimetype='application/pdf',
                     as_attachment=False,
                     download_name=fname)
