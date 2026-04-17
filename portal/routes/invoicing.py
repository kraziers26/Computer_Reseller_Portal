from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from ..auth_utils import require_role
from ..db import db_cursor
import uuid, io
from datetime import date

invoicing_bp = Blueprint('invoicing', __name__, url_prefix='/invoicing')

COMPANY_DATA = {
    'SE': {
        'name': 'Sunny Enterprise Corp',
        'addr': '9400 W Flagler ST\nMiami FL 33174',
        'beneficiario': 'Sunny Esnug Enterprise\n9400 W Flagler ST\nMiami FL 33174',
        'banco': 'Chase Bank', 'route': '267084131', 'account': '890615856',
        'payable': 'Sunny Esnug Enterprise', 'seq': 'invoice_seq_se',
    },
    'MS': {
        'name': 'Medara Studio',
        'addr': '10739 NW 70th Lane\nDoral FL 33178',
        'beneficiario': 'Medara Corp\n10739 NW 70th Lane\nDoral FL 33178',
        'banco': 'Chase Bank', 'route': '267084131', 'account': '795651980',
        'payable': 'Medara Corp', 'seq': 'invoice_seq_ms',
    },
}

MIN_ITEM_PRICE = 1.0  # Items under $1 are omitted from invoices


def get_next_invoice_number(cur, code):
    # Check reusable pool first
    cur.execute("""
        SELECT inv_number FROM invoice_number_pool
        WHERE company_code=%s ORDER BY recycled_at LIMIT 1
    """, (code,))
    row = cur.fetchone()
    if row:
        num = row['inv_number']
        cur.execute("DELETE FROM invoice_number_pool WHERE inv_number=%s AND company_code=%s",
                    (num, code))
        return num
    seq = COMPANY_DATA[code]['seq']
    cur.execute(f"SELECT nextval('{seq}') AS n")
    n = cur.fetchone()['n']
    return f"{int(n):05d}-{code}"


def _load_received_orders(cur):
    """Load received orders with their line items (items >= $1 only)."""
    cur.execute("""
        SELECT t.transaction_id, t.order_number, t.retailer, t.purchase_date,
               t.price_total, u.username AS person_name,
               c.company_name, c.company_id,
               c.company_short_code
        FROM transactions t
        LEFT JOIN dim_users u      ON t.user_id = u.user_id
        LEFT JOIN dim_companies c  ON t.company_id = c.company_id
        WHERE t.fulfillment_status='received' AND t.is_active=TRUE
        ORDER BY t.retailer, t.order_number
    """)
    orders = cur.fetchall()

    # Load line items per order, filtering out < $1 items
    order_items = {}
    if orders:
        txn_ids = [str(o['transaction_id']) for o in orders]
        cur.execute("""
            SELECT transaction_id, item_id, item_description,
                   sku_model_color, quantity, unit_price, line_total
            FROM transaction_items
            WHERE transaction_id = ANY(%s::uuid[])
              AND unit_price >= %s
            ORDER BY transaction_id, item_description
        """, (txn_ids, MIN_ITEM_PRICE))
        for item in cur.fetchall():
            tid = str(item['transaction_id'])
            order_items.setdefault(tid, []).append(item)

    return orders, order_items


# ── Invoice History ───────────────────────────────────────────────────────────

@invoicing_bp.route('/')
@login_required
@require_role('admin')
def index():
    f_company  = request.args.get('company', type=int)
    f_customer = request.args.get('customer')
    f_status   = request.args.get('status', '')
    f_from     = request.args.get('date_from', '')
    f_to       = request.args.get('date_to', '')

    conditions = ["1=1"]
    params = []
    if f_company:  conditions.append("i.company_id=%s");        params.append(f_company)
    if f_customer: conditions.append("i.customer_id=%s::uuid"); params.append(f_customer)
    if f_status:   conditions.append("i.status=%s");             params.append(f_status)
    if f_from:     conditions.append("i.invoice_date>=%s");      params.append(f_from)
    if f_to:       conditions.append("i.invoice_date<=%s");      params.append(f_to)
    where = ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT i.invoice_id, i.invoice_number, i.invoice_date, i.status,
                   i.subtotal, i.total, i.created_at,
                   c.company_name, c.company_short_code,
                   cu.customer_name,
                   u.username AS created_by,
                   COUNT(ii.item_id) AS line_count
            FROM invoices i
            LEFT JOIN dim_companies  c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers  cu ON i.customer_id = cu.customer_id
            LEFT JOIN dim_users      u  ON i.created_by  = u.user_id
            LEFT JOIN invoice_items  ii ON i.invoice_id  = ii.invoice_id
            WHERE {where}
            GROUP BY i.invoice_id, i.invoice_number, i.invoice_date, i.status,
                     i.subtotal, i.total, i.created_at,
                     c.company_name, c.company_short_code,
                     cu.customer_name, u.username
            ORDER BY i.created_at DESC
        """, params)
        invoices = cur.fetchall()

        cur.execute("SELECT company_id, company_name, company_short_code FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT customer_id, customer_name FROM dim_customers WHERE is_active=TRUE ORDER BY customer_name")
        customers = cur.fetchall()

        cur.execute("""
            SELECT
                COUNT(*) FILTER (WHERE status='draft') AS drafts,
                COUNT(*) FILTER (WHERE status='sent')  AS sent,
                COUNT(*) FILTER (WHERE status='paid')  AS paid,
                COALESCE(SUM(total) FILTER (WHERE status='sent'), 0)  AS pending_amount,
                COALESCE(SUM(total) FILTER (WHERE status='paid'), 0)  AS paid_amount
            FROM invoices
        """)
        kpis = cur.fetchone()

    return render_template('invoicing/index.html',
                           invoices=invoices, companies=companies, customers=customers,
                           kpis=kpis,
                           filters={'company':f_company,'customer':f_customer,
                                    'status':f_status,'date_from':f_from,'date_to':f_to})


# ── New Invoice ───────────────────────────────────────────────────────────────

@invoicing_bp.route('/new', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def new_invoice():
    if request.method == 'POST':
        return _save_invoice(existing_id=None)

    with db_cursor() as (cur, _):
        orders, order_items = _load_received_orders(cur)
        cur.execute("SELECT company_id, company_name, company_short_code FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT customer_id, customer_name FROM dim_customers WHERE is_active=TRUE ORDER BY customer_name")
        customers = cur.fetchall()
        retailers = sorted(set(o['retailer'] for o in orders if o['retailer']))
        persons   = sorted(set(o['person_name'] for o in orders if o['person_name']))

    return render_template('invoicing/new.html',
                           orders=orders, order_items=order_items,
                           companies=companies, customers=customers,
                           retailers=retailers, persons=persons,
                           today=str(date.today()),
                           edit_mode=False, existing_invoice=None,
                           existing_txn_ids=[])


# ── Edit Invoice ──────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def edit_invoice(invoice_id):
    sid = str(invoice_id)

    if request.method == 'POST':
        return _save_invoice(existing_id=sid)

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code, cu.customer_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            WHERE i.invoice_id=%s AND i.status='draft'
        """, (sid,))
        inv = cur.fetchone()
        if not inv:
            flash('Invoice not found or not editable.', 'error')
            return redirect(url_for('invoicing.index'))

        # Current orders on this invoice
        cur.execute("""
            SELECT DISTINCT transaction_id::text FROM invoice_items
            WHERE invoice_id=%s
        """, (sid,))
        existing_txn_ids = [r['transaction_id'] for r in cur.fetchall()]

        orders, order_items = _load_received_orders(cur)
        cur.execute("SELECT company_id, company_name, company_short_code FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT customer_id, customer_name FROM dim_customers WHERE is_active=TRUE ORDER BY customer_name")
        customers = cur.fetchall()
        retailers = sorted(set(o['retailer'] for o in orders if o['retailer']))
        persons   = sorted(set(o['person_name'] for o in orders if o['person_name']))

    return render_template('invoicing/new.html',
                           orders=orders, order_items=order_items,
                           companies=companies, customers=customers,
                           retailers=retailers, persons=persons,
                           today=str(date.today()),
                           edit_mode=True, existing_invoice=inv,
                           existing_txn_ids=existing_txn_ids)


# ── Shared save logic ─────────────────────────────────────────────────────────

def _save_invoice(existing_id=None):
    company_id   = request.form.get('company_id', type=int)
    customer_id  = request.form.get('customer_id', '').strip()
    new_customer = request.form.get('new_customer_name', '').strip()
    markup_pct   = request.form.get('batch_markup', type=float) or 1.0
    other_amount = request.form.get('other_amount', type=float) or 0.0
    other_label  = request.form.get('other_label', '').strip()
    invoice_date = request.form.get('invoice_date') or str(date.today())
    txn_ids      = request.form.getlist('txn_ids')

    if not txn_ids:
        flash('Select at least one order.', 'error')
        return redirect(url_for('invoicing.edit_invoice', invoice_id=existing_id)
                        if existing_id else url_for('invoicing.new_invoice'))

    with db_cursor() as (cur, conn):
        cur.execute("SELECT company_short_code FROM dim_companies WHERE company_id=%s", (company_id,))
        row = cur.fetchone()
        code = (row['company_short_code'] if row else 'SE').upper()
        if code not in COMPANY_DATA:
            code = 'SE'

        if new_customer and not customer_id:
            customer_id = str(uuid.uuid4())
            cur.execute("INSERT INTO dim_customers (customer_id, customer_name) VALUES (%s,%s)",
                        (customer_id, new_customer))

        if existing_id:
            # Edit: find orders that were removed → release back to received
            cur.execute("SELECT DISTINCT transaction_id::text FROM invoice_items WHERE invoice_id=%s",
                        (existing_id,))
            old_ids = set(r['transaction_id'] for r in cur.fetchall())
            new_ids = set(txn_ids)
            released = old_ids - new_ids
            if released:
                cur.execute("""
                    UPDATE transactions SET fulfillment_status='received',
                        fulfillment_status_updated_at=NOW()
                    WHERE transaction_id = ANY(%s::uuid[])
                """, (list(released),))
            # Delete old items
            cur.execute("DELETE FROM invoice_items WHERE invoice_id=%s", (existing_id,))
            invoice_id  = existing_id
            inv_number  = request.form.get('invoice_number')
        else:
            invoice_id = str(uuid.uuid4())
            inv_number = get_next_invoice_number(cur, code)

        # Build line items — skip < $1, consolidate duplicates by description+price
        cur.execute("""
            SELECT t.transaction_id, t.order_number, t.retailer,
                   ti.item_description, ti.sku_model_color,
                   ti.quantity, ti.unit_price
            FROM transactions t
            LEFT JOIN transaction_items ti ON t.transaction_id = ti.transaction_id
            WHERE t.transaction_id = ANY(%s::uuid[]) AND t.is_active=TRUE
              AND ti.unit_price >= %s
            ORDER BY ti.item_description, ti.unit_price
        """, (txn_ids, MIN_ITEM_PRICE))
        raw_items = cur.fetchall()

        # Consolidate: same description + same unit_price → merge qty, keep first transaction_id
        from collections import OrderedDict
        merged = OrderedDict()
        for item in raw_items:
            desc  = (item['item_description'] or item['retailer'] or '').strip()
            price = float(item['unit_price'] or 0)
            key   = (desc.lower(), price)
            if key in merged:
                merged[key]['qty'] += int(item['quantity'] or 1)
            else:
                merged[key] = {
                    'transaction_id': str(item['transaction_id']),
                    'description':    desc,
                    'sku':            item['sku_model_color'] or '',
                    'qty':            int(item['quantity'] or 1),
                    'unit_cost':      price,
                }

        subtotal = 0.0
        line_items = []
        for i, (key, item) in enumerate(merged.items()):
            item_markup = float(request.form.get(f'markup_{i}', markup_pct))
            unit_cost   = item['unit_cost']
            unit_price  = round(unit_cost * (1 + item_markup / 100), 2)
            qty         = item['qty']
            line_total  = round(unit_price * qty, 2)
            subtotal   += line_total
            line_items.append({
                'item_id':        str(uuid.uuid4()),
                'transaction_id': item['transaction_id'],
                'description':    item['description'],
                'sku':            item['sku'],
                'qty':            qty,
                'unit_cost':      unit_cost,
                'markup_pct':     item_markup,
                'unit_price':     unit_price,
                'line_total':     line_total,
                'sort_order':     i,
            })

        subtotal = round(subtotal, 2)
        total    = round(subtotal + other_amount, 2)

        if existing_id:
            cur.execute("""
                UPDATE invoices SET company_id=%s, customer_id=%s, invoice_date=%s,
                    batch_markup_pct=%s, subtotal=%s, other_amount=%s,
                    other_label=%s, total=%s
                WHERE invoice_id=%s
            """, (company_id, customer_id or None, invoice_date,
                  markup_pct, subtotal, other_amount,
                  other_label or None, total, invoice_id))
        else:
            cur.execute("""
                INSERT INTO invoices (invoice_id, invoice_number, company_id, customer_id,
                    created_by, invoice_date, batch_markup_pct,
                    subtotal, other_amount, other_label, total, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'draft')
            """, (invoice_id, inv_number, company_id,
                  customer_id or None, current_user.id,
                  invoice_date, markup_pct,
                  subtotal, other_amount, other_label or None, total))

        for li in line_items:
            cur.execute("""
                INSERT INTO invoice_items
                    (item_id, invoice_id, transaction_id, item_description, sku,
                     quantity, unit_cost, markup_pct, unit_price, line_total, sort_order)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (li['item_id'], invoice_id, li['transaction_id'],
                  li['description'], li['sku'], li['qty'],
                  li['unit_cost'], li['markup_pct'], li['unit_price'],
                  li['line_total'], li['sort_order']))

        # Mark all selected transactions as invoiced
        cur.execute("""
            UPDATE transactions SET fulfillment_status='invoiced',
                fulfillment_status_updated_at=NOW()
            WHERE transaction_id = ANY(%s::uuid[])
        """, (txn_ids,))

    action = 'updated' if existing_id else 'created'
    flash(f'Invoice {inv_number} {action}!', 'success')
    return redirect(url_for('invoicing.view_invoice', invoice_id=invoice_id))


# ── Delete Invoice (Draft only) ───────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/delete', methods=['POST'])
@login_required
@require_role('admin')
def delete_invoice(invoice_id):
    sid = str(invoice_id)
    with db_cursor() as (cur, conn):
        cur.execute("SELECT invoice_number, status, company_id FROM invoices WHERE invoice_id=%s", (sid,))
        inv = cur.fetchone()
        if not inv:
            flash('Invoice not found.', 'error')
            return redirect(url_for('invoicing.index'))
        if inv['status'] != 'draft':
            flash('Only Draft invoices can be deleted.', 'error')
            return redirect(url_for('invoicing.index'))

        # Release orders back to received
        cur.execute("""
            UPDATE transactions t
            SET fulfillment_status='received', fulfillment_status_updated_at=NOW()
            FROM invoice_items ii
            WHERE ii.transaction_id = t.transaction_id AND ii.invoice_id=%s
        """, (sid,))

        # Recycle invoice number
        cur.execute("SELECT company_short_code FROM dim_companies WHERE company_id=%s",
                    (inv['company_id'],))
        co = cur.fetchone()
        if co:
            cur.execute("""
                INSERT INTO invoice_number_pool (inv_number, company_code)
                VALUES (%s, %s)
                ON CONFLICT DO NOTHING
            """, (inv['invoice_number'], co['company_short_code']))

        # Delete invoice (cascades to invoice_items)
        cur.execute("DELETE FROM invoices WHERE invoice_id=%s", (sid,))

    flash(f'Invoice {inv["invoice_number"]} deleted. Orders returned to Received pool.', 'success')
    return redirect(url_for('invoicing.index'))


# ── Remove single item from draft invoice ────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/remove-item', methods=['POST'])
@login_required
@require_role('admin')
def remove_item(invoice_id):
    sid = str(invoice_id)
    data = request.get_json()
    item_id = data.get('item_id')
    if not item_id:
        return jsonify({'error': 'Missing item_id'}), 400

    with db_cursor() as (cur, conn):
        cur.execute("SELECT status FROM invoices WHERE invoice_id=%s", (sid,))
        inv = cur.fetchone()
        if not inv or inv['status'] != 'draft':
            return jsonify({'error': 'Invoice not editable'}), 400

        cur.execute(
            "SELECT transaction_id FROM invoice_items WHERE item_id=%s AND invoice_id=%s",
            (item_id, sid))
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Item not found'}), 404
        txn_id = str(row['transaction_id'])

        cur.execute("DELETE FROM invoice_items WHERE item_id=%s", (item_id,))

        cur.execute(
            "SELECT COUNT(*) AS n FROM invoice_items "
            "WHERE invoice_id=%s AND transaction_id=%s::uuid",
            (sid, txn_id))
        remaining = cur.fetchone()['n']
        released_order = False
        if remaining == 0:
            cur.execute(
                "UPDATE transactions SET fulfillment_status='received',"
                "fulfillment_status_updated_at=NOW() "
                "WHERE transaction_id=%s::uuid",
                (txn_id,))
            released_order = True

        cur.execute(
            "SELECT COALESCE(SUM(line_total),0) AS s FROM invoice_items WHERE invoice_id=%s",
            (sid,))
        new_subtotal = float(cur.fetchone()['s'])
        cur.execute("SELECT other_amount FROM invoices WHERE invoice_id=%s", (sid,))
        other = float(cur.fetchone()['other_amount'] or 0)
        new_total = round(new_subtotal + other, 2)
        cur.execute(
            "UPDATE invoices SET subtotal=%s, total=%s WHERE invoice_id=%s",
            (new_subtotal, new_total, sid))

    return jsonify({
        'ok': True,
        'subtotal': f"{new_subtotal:,.2f}",
        'total': f"{new_total:,.2f}",
        'removed_order': released_order
    })


# ── Invoice Detail ────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>')
@login_required
@require_role('admin')
def view_invoice(invoice_id):
    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code,
                   cu.customer_name, u.username AS created_by_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            LEFT JOIN dim_users u      ON i.created_by  = u.user_id
            WHERE i.invoice_id=%s
        """, (str(invoice_id),))
        inv = cur.fetchone()
        if not inv:
            flash('Invoice not found.', 'error')
            return redirect(url_for('invoicing.index'))

        cur.execute("SELECT * FROM invoice_items WHERE invoice_id=%s ORDER BY sort_order",
                    (str(invoice_id),))
        items = cur.fetchall()

    code = (inv['company_short_code'] or 'SE').upper()
    co   = COMPANY_DATA.get(code, COMPANY_DATA['SE'])
    return render_template('invoicing/view.html', inv=inv, items=items, co=co)


# ── Update Status ─────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/status', methods=['POST'])
@login_required
@require_role('admin')
def update_status(invoice_id):
    status = request.form.get('status')
    if status in ('draft', 'sent', 'paid'):
        with db_cursor() as (cur, conn):
            cur.execute("UPDATE invoices SET status=%s WHERE invoice_id=%s",
                        (status, str(invoice_id)))
    return redirect(url_for('invoicing.view_invoice', invoice_id=str(invoice_id)))


# ── Update item description (draft only) ─────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/update-item', methods=['POST'])
@login_required
@require_role('admin')
def update_item_description(invoice_id):
    sid = str(invoice_id)
    data = request.get_json()
    item_id = data.get('item_id')
    new_desc = (data.get('description') or '').strip()
    if not item_id or not new_desc:
        return jsonify({'error': 'Missing fields'}), 400
    with db_cursor() as (cur, conn):
        cur.execute("SELECT status FROM invoices WHERE invoice_id=%s", (sid,))
        inv = cur.fetchone()
        if not inv or inv['status'] != 'draft':
            return jsonify({'error': 'Not editable'}), 400
        cur.execute(
            "UPDATE invoice_items SET item_description=%s WHERE item_id=%s AND invoice_id=%s",
            (new_desc, item_id, sid))
    return jsonify({'ok': True, 'description': new_desc})


# ── Export Excel ──────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/export-excel')
@login_required
@require_role('admin')
def export_excel(invoice_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code, cu.customer_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            WHERE i.invoice_id=%s
        """, (str(invoice_id),))
        inv = cur.fetchone()
        cur.execute("SELECT * FROM invoice_items WHERE invoice_id=%s ORDER BY sort_order",
                    (str(invoice_id),))
        items = cur.fetchall()

    code = (inv['company_short_code'] or 'SE').upper()
    co   = COMPANY_DATA.get(code, COMPANY_DATA['SE'])

    wb = Workbook()
    ws = wb.active
    ws.title = inv['invoice_number']
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_margins.left  = 0.6
    ws.page_margins.right = 0.6

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 2

    DARK  = '1A1A2E'
    ACCENT= '4F6EF7'
    LIGHT = 'F0F2FF'
    GRAY  = 'E8E9F0'
    WHITE = 'FFFFFF'
    MUTED = '6B7280'

    def c(row, col, val='', bold=False, sz=10, fg=None, bg=None,
           align='left', valign='center', wrap=False, fmt=None, italic=False):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='Calibri', bold=bold, size=sz,
                         color=fg or '000000', italic=italic)
        cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
        if bg:
            cell.fill = PatternFill('solid', fgColor=bg)
        if fmt:
            cell.number_format = fmt
        return cell

    def fill_row(row, c1, c2, bg):
        for col in range(c1, c2+1):
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=bg)

    def merge(row, c1, c2, val='', **kw):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        return c(row, c1, val, **kw)

    r = 1

    # Accent top bar
    ws.row_dimensions[r].height = 5
    fill_row(r, 1, 6, ACCENT); r += 1

    # Company name
    ws.row_dimensions[r].height = 32
    fill_row(r, 1, 6, DARK)
    merge(r, 2, 3, co['name'], bold=True, sz=16, fg=WHITE, bg=DARK, valign='center')
    inv_num = "INVOICE #" + inv['invoice_number']
    merge(r, 4, 5, inv_num, bold=True, sz=13, fg=WHITE, bg=DARK,
          align='right', valign='center')
    r += 1

    # Address + date
    ws.row_dimensions[r].height = 42
    fill_row(r, 1, 6, DARK)
    addr_val = co['addr']
    merge(r, 2, 3, addr_val, sz=9, fg='9CA3AF', bg=DARK, wrap=True, valign='top')
    dt = inv['invoice_date']
    date_str = dt.strftime('%B %d, %Y') if hasattr(dt, 'strftime') else str(dt)
    merge(r, 4, 5, date_str, sz=10, fg='CBD5E1', bg=DARK, align='right', valign='top')
    r += 1

    # Accent bottom bar
    ws.row_dimensions[r].height = 4
    fill_row(r, 1, 6, ACCENT); r += 1

    # Spacer
    ws.row_dimensions[r].height = 12; r += 1

    # Bill to
    ws.row_dimensions[r].height = 13
    merge(r, 2, 5, 'BILL TO', bold=True, sz=8, fg=MUTED)
    r += 1
    ws.row_dimensions[r].height = 20
    merge(r, 2, 5, inv['customer_name'] or '—', bold=True, sz=12)
    r += 1

    # Spacer
    ws.row_dimensions[r].height = 10; r += 1

    # Items header
    ws.row_dimensions[r].height = 22
    fill_row(r, 1, 6, ACCENT)
    for col, txt, al in [(2,'DESCRIPTION','left'),(3,'PRICE','right'),
                          (4,'QTY','center'),(5,'TOTAL','right')]:
        c(r, col, txt, bold=True, sz=9, fg=WHITE, bg=ACCENT, align=al)
    r += 1

    for idx, item in enumerate(items):
        bg = WHITE if idx % 2 == 0 else LIGHT
        ws.row_dimensions[r].height = 34
        fill_row(r, 1, 6, bg)
        bdr = Border(bottom=Side(style='thin', color=GRAY))
        desc_cell = ws.cell(row=r, column=2, value=item['item_description'] or '')
        desc_cell.font = Font(name='Calibri', size=10)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
        desc_cell.fill = PatternFill('solid', fgColor=bg)
        desc_cell.border = bdr

        pc = ws.cell(row=r, column=3, value=float(item['unit_price']))
        pc.font = Font(name='Calibri', size=10)
        pc.number_format = '"$"#,##0.00'
        pc.alignment = Alignment(horizontal='right', vertical='center')
        pc.fill = PatternFill('solid', fgColor=bg)
        pc.border = bdr

        qc = ws.cell(row=r, column=4, value=int(item['quantity']))
        qc.font = Font(name='Calibri', size=10)
        qc.alignment = Alignment(horizontal='center', vertical='center')
        qc.fill = PatternFill('solid', fgColor=bg)
        qc.border = bdr

        tc = ws.cell(row=r, column=5, value=float(item['line_total']))
        tc.font = Font(name='Calibri', bold=True, size=10)
        tc.number_format = '"$"#,##0.00'
        tc.alignment = Alignment(horizontal='right', vertical='center')
        tc.fill = PatternFill('solid', fgColor=bg)
        tc.border = bdr
        r += 1

    # Spacer
    ws.row_dimensions[r].height = 8; r += 1

    # Subtotal / Tax / Other
    for lbl, val, bold in [
        ('SUBTOTAL', float(inv['subtotal']), False),
        ('TAX RATE', '0.00%', False),
        ((inv.get('other_label') or 'OTHER').upper(), float(inv['other_amount'] or 0), False),
    ]:
        ws.row_dimensions[r].height = 18
        c(r, 4, lbl, sz=9, fg=MUTED, align='right')
        vc = ws.cell(row=r, column=5, value=val)
        vc.font = Font(name='Calibri', size=10, color=MUTED)
        vc.alignment = Alignment(horizontal='right', vertical='center')
        if isinstance(val, float):
            vc.number_format = '"$"#,##0.00'
        r += 1

    # Total bar
    ws.row_dimensions[r].height = 28
    fill_row(r, 1, 6, DARK)
    c(r, 4, 'TOTAL', bold=True, sz=11, fg=WHITE, bg=DARK, align='right')
    tc = ws.cell(row=r, column=5, value=float(inv['total']))
    tc.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    tc.number_format = '"$"#,##0.00'
    tc.alignment = Alignment(horizontal='right', vertical='center')
    tc.fill = PatternFill('solid', fgColor=DARK)
    r += 1

    # Spacer
    ws.row_dimensions[r].height = 16; r += 1

    # Thank you
    ws.row_dimensions[r].height = 20
    merge(r, 2, 5, 'THANK YOU FOR YOUR BUSINESS!',
          bold=True, sz=10, fg=ACCENT, align='center')
    r += 1

    # Thin divider
    ws.row_dimensions[r].height = 3
    fill_row(r, 2, 5, GRAY); r += 1

    # Spacer
    ws.row_dimensions[r].height = 12; r += 1

    # Banking section
    ws.row_dimensions[r].height = 13
    merge(r, 2, 5, 'TERMS & CONDITIONS', bold=True, sz=8, fg=MUTED)
    r += 1

    bene_lines  = co['beneficiario'].split('\n')
    banco_lines = [co['banco'],
                   'Route number: ' + co['route'],
                   'Account number: ' + co['account']]

    c(r, 2, 'Beneficiario:', bold=True, sz=9)
    c(r, 4, 'Banco:', bold=True, sz=9)
    r += 1

    max_lines = max(len(bene_lines), len(banco_lines))
    for i in range(max_lines):
        ws.row_dimensions[r].height = 15
        if i < len(bene_lines):
            c(r, 2, bene_lines[i], sz=9, fg='374151')
        if i < len(banco_lines):
            c(r, 4, banco_lines[i], sz=9, fg='374151')
        r += 1

    r += 1
    ws.row_dimensions[r].height = 14
    merge(r, 2, 5, 'Make all checks payable to ' + co['payable'],
          sz=9, fg=MUTED, italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Invoice_' + inv['invoice_number'] + '.xlsx')



# ── Export PDF ────────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/export-pdf')
@login_required
@require_role('admin')
def export_pdf(invoice_id):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code, cu.customer_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            WHERE i.invoice_id=%s
        """, (str(invoice_id),))
        inv = cur.fetchone()
        cur.execute("SELECT * FROM invoice_items WHERE invoice_id=%s ORDER BY sort_order",
                    (str(invoice_id),))
        items = cur.fetchall()

    code = (inv['company_short_code'] or 'SE').upper()
    co   = COMPANY_DATA.get(code, COMPANY_DATA['SE'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)

    normal = ParagraphStyle('n', fontName='Helvetica', fontSize=10, leading=14)
    bold   = ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=10, leading=14)
    right  = ParagraphStyle('r', fontName='Helvetica', fontSize=10, leading=14, alignment=TA_RIGHT)
    rbold  = ParagraphStyle('rb', fontName='Helvetica-Bold', fontSize=10, leading=14, alignment=TA_RIGHT)
    center = ParagraphStyle('c', fontName='Helvetica', fontSize=10, leading=14, alignment=TA_CENTER)
    title  = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=16, leading=20)
    small  = ParagraphStyle('s', fontName='Helvetica', fontSize=9, leading=12, textColor=colors.HexColor('#666666'))

    story = []
    addr_html = co['addr'].replace('\n','<br/>')
    inv_date  = inv['invoice_date']
    date_str  = inv_date.strftime('%m/%d/%y') if hasattr(inv_date,'strftime') else str(inv_date)

    ht = Table([
        [Paragraph(co['name'], title), Paragraph(f"<b>INVOICE #{inv['invoice_number']}</b>", rbold)],
        [Paragraph(addr_html, small),  Paragraph(date_str, right)],
    ], colWidths=[4*inch, 3*inch])
    ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    story += [ht, Spacer(1,0.2*inch), Paragraph('To', small),
              Paragraph(inv['customer_name'] or '—', bold), Spacer(1,0.25*inch)]

    table_data = [[Paragraph('Description',bold),Paragraph('Price',rbold),
                   Paragraph('QTY',rbold),Paragraph('Total',rbold)]]
    for item in items:
        table_data.append([
            Paragraph(item['item_description'] or '', normal),
            Paragraph(f"${float(item['unit_price']):,.2f}", right),
            Paragraph(str(item['quantity']), right),
            Paragraph(f"${float(item['line_total']):,.2f}", right),
        ])
    t = Table(table_data, colWidths=[3.8*inch,1.2*inch,0.7*inch,1.3*inch], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#EEEEEE')),
        ('LINEBELOW',(0,0),(-1,0),0.5,colors.black),
        ('LINEBELOW',(0,1),(-1,-1),0.25,colors.HexColor('#DDDDDD')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story += [t, Spacer(1,0.15*inch)]

    other_lbl = inv.get('other_label') or 'OTHER'
    tt = Table([
        [Paragraph('SUBTOTAL',right), Paragraph(f"${float(inv['subtotal']):,.2f}",right)],
        [Paragraph('TAX RATE',right), Paragraph('0.00%',right)],
        [Paragraph(other_lbl,right),  Paragraph(f"${float(inv['other_amount'] or 0):,.2f}",right)],
        [Paragraph('<b>TOTAL</b>',rbold), Paragraph(f"<b>${float(inv['total']):,.2f}</b>",rbold)],
    ], colWidths=[5.5*inch,1.5*inch])
    tt.setStyle(TableStyle([
        ('LINEABOVE',(0,3),(-1,3),0.5,colors.black),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
    ]))
    story += [tt, Spacer(1,0.3*inch), Paragraph('THANK YOU FOR YOUR BUSINESS!',center),
              Spacer(1,0.25*inch), Paragraph('<b>TERMS &amp; CONDITIONS</b>',bold),
              Spacer(1,0.08*inch)]

    bene_html = co['beneficiario'].replace('\n','<br/>')
    bt = Table([
        [Paragraph('<b>Beneficiario:</b>',bold), Paragraph('<b>Banco:</b>',bold)],
        [Paragraph(bene_html,normal),
         Paragraph(f"{co['banco']}<br/>Route number: {co['route']}<br/>Account number: {co['account']}",normal)],
    ], colWidths=[3.5*inch,3.5*inch])
    bt.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),
                             ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3)]))
    story += [bt, Spacer(1,0.1*inch),
              Paragraph(f"Make all checks payable to {co['payable']}", small)]

    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"Invoice_{inv['invoice_number']}.pdf")
