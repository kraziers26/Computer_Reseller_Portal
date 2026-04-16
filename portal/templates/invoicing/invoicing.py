from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from ..auth_utils import require_role
from ..db import db_cursor
import uuid, io
from datetime import date

invoicing_bp = Blueprint('invoicing', __name__, url_prefix='/invoicing')

COMPANY_DATA = {
    'SE': {
        'name': 'Sunny Enterprise Corp',
        'addr': '9400 W Flagler ST\nMiami FL 33174',
        'beneficiario': 'Sunny Esnug Enterprise\n9400 W Flagler ST\nMiami FL 33174',
        'banco': 'Chase Bank',
        'route': '267084131',
        'account': '890615856',
        'payable': 'Sunny Esnug Enterprise',
        'seq': 'invoice_seq_se',
    },
    'MS': {
        'name': 'Medara Studio',
        'addr': '10739 NW 70th Lane\nDoral FL 33178',
        'beneficiario': 'Medara Corp\n10739 NW 70th Lane\nDoral FL 33178',
        'banco': 'Chase Bank',
        'route': '267084131',
        'account': '795651980',
        'payable': 'Medara Corp',
        'seq': 'invoice_seq_ms',
    },
}

def get_next_invoice_number(cur, code):
    seq = COMPANY_DATA[code]['seq']
    cur.execute(f"SELECT nextval('{seq}') AS n")
    n = cur.fetchone()['n']
    return f"{int(n):05d}-{code}"


# ── Invoice History ───────────────────────────────────────────────────────────

@invoicing_bp.route('/')
@login_required
@require_role('admin')
def index():
    f_company  = request.args.get('company', type=int)
    f_customer = request.args.get('customer')
    f_status   = request.args.get('status', '')
    f_from     = request.args.get('date_from', '')
    f_to       = request.args.get('date_to', '')

    conditions = ["1=1"]
    params = []
    if f_company:  conditions.append("i.company_id=%s");       params.append(f_company)
    if f_customer: conditions.append("i.customer_id=%s::uuid"); params.append(f_customer)
    if f_status:   conditions.append("i.status=%s");            params.append(f_status)
    if f_from:     conditions.append("i.invoice_date>=%s");     params.append(f_from)
    if f_to:       conditions.append("i.invoice_date<=%s");     params.append(f_to)
    where = ' AND '.join(conditions)

    with db_cursor() as (cur, _):
        cur.execute(f"""
            SELECT i.invoice_id, i.invoice_number, i.invoice_date, i.status,
                   i.subtotal, i.total, i.created_at,
                   c.company_name, c.company_short_code,
                   cu.customer_name,
                   u.username AS created_by,
                   COUNT(ii.item_id) AS line_count
            FROM invoices i
            LEFT JOIN dim_companies  c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers  cu ON i.customer_id = cu.customer_id
            LEFT JOIN dim_users      u  ON i.created_by  = u.user_id
            LEFT JOIN invoice_items  ii ON i.invoice_id  = ii.invoice_id
            WHERE {where}
            GROUP BY i.invoice_id, i.invoice_number, i.invoice_date, i.status,
                     i.subtotal, i.total, i.created_at,
                     c.company_name, c.company_short_code,
                     cu.customer_name, u.username
            ORDER BY i.created_at DESC
        """, params)
        invoices = cur.fetchall()

        cur.execute("SELECT company_id, company_name, company_short_code FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT customer_id, customer_name FROM dim_customers WHERE is_active=TRUE ORDER BY customer_name")
        customers = cur.fetchall()

        # Totals for KPIs
        cur.execute("""
            SELECT
                COUNT(*) FILTER (WHERE status='draft') AS drafts,
                COUNT(*) FILTER (WHERE status='sent')  AS sent,
                COUNT(*) FILTER (WHERE status='paid')  AS paid,
                COALESCE(SUM(total) FILTER (WHERE status='sent'), 0)  AS pending_amount,
                COALESCE(SUM(total) FILTER (WHERE status='paid'), 0)  AS paid_amount
            FROM invoices
        """)
        kpis = cur.fetchone()

    return render_template('invoicing/index.html',
                           invoices=invoices, companies=companies, customers=customers,
                           kpis=kpis,
                           filters={'company':f_company,'customer':f_customer,
                                    'status':f_status,'date_from':f_from,'date_to':f_to})


# ── New Invoice Builder ───────────────────────────────────────────────────────

@invoicing_bp.route('/new', methods=['GET', 'POST'])
@login_required
@require_role('admin')
def new_invoice():
    if request.method == 'POST':
        company_id    = request.form.get('company_id', type=int)
        customer_id   = request.form.get('customer_id', '').strip()
        new_customer  = request.form.get('new_customer_name', '').strip()
        markup_pct    = request.form.get('batch_markup', type=float) or 1.0
        other_amount  = request.form.get('other_amount', type=float) or 0.0
        other_label   = request.form.get('other_label', '').strip()
        invoice_date  = request.form.get('invoice_date') or str(date.today())
        txn_ids       = request.form.getlist('txn_ids')

        if not txn_ids:
            flash('Select at least one order to invoice.', 'error')
            return redirect(url_for('invoicing.new_invoice'))

        # Get company short code
        with db_cursor() as (cur, conn):
            cur.execute("SELECT company_short_code FROM dim_companies WHERE company_id=%s", (company_id,))
            row = cur.fetchone()
            code = (row['company_short_code'] if row else 'SE').upper()
            if code not in COMPANY_DATA:
                code = 'SE'

            # Create customer if new
            if new_customer and not customer_id:
                customer_id = str(uuid.uuid4())
                cur.execute("INSERT INTO dim_customers (customer_id, customer_name) VALUES (%s,%s)",
                            (customer_id, new_customer))

            # Generate invoice number
            inv_number = get_next_invoice_number(cur, code)
            invoice_id = str(uuid.uuid4())

            # Build line items from transaction_items
            cur.execute("""
                SELECT t.transaction_id, t.order_number, t.retailer,
                       ti.item_description, ti.sku_model_color,
                       ti.quantity, ti.unit_price
                FROM transactions t
                LEFT JOIN transaction_items ti ON t.transaction_id = ti.transaction_id
                WHERE t.transaction_id = ANY(%s::uuid[]) AND t.is_active=TRUE
                ORDER BY t.order_number, ti.item_description
            """, (txn_ids,))
            raw_items = cur.fetchall()

            # Compute per-item overrides from form
            subtotal = 0.0
            line_items = []
            for i, item in enumerate(raw_items):
                item_markup = float(request.form.get(f'markup_{i}', markup_pct))
                unit_cost   = float(item['unit_price'] or 0)
                unit_price  = round(unit_cost * (1 + item_markup / 100), 2)
                qty         = int(item['quantity'] or 1)
                line_total  = round(unit_price * qty, 2)
                subtotal   += line_total
                line_items.append({
                    'item_id':      str(uuid.uuid4()),
                    'transaction_id': str(item['transaction_id']),
                    'description':  item['item_description'] or item['retailer'] or '',
                    'sku':          item['sku_model_color'] or '',
                    'qty':          qty,
                    'unit_cost':    unit_cost,
                    'markup_pct':   item_markup,
                    'unit_price':   unit_price,
                    'line_total':   line_total,
                    'sort_order':   i,
                })

            subtotal = round(subtotal, 2)
            total    = round(subtotal + other_amount, 2)

            # Save invoice
            cur.execute("""
                INSERT INTO invoices (invoice_id, invoice_number, company_id, customer_id,
                    created_by, invoice_date, batch_markup_pct,
                    subtotal, other_amount, other_label, total, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'draft')
            """, (invoice_id, inv_number, company_id,
                  customer_id or None, current_user.id,
                  invoice_date, markup_pct,
                  subtotal, other_amount, other_label or None, total))

            for li in line_items:
                cur.execute("""
                    INSERT INTO invoice_items
                        (item_id, invoice_id, transaction_id, item_description, sku,
                         quantity, unit_cost, markup_pct, unit_price, line_total, sort_order)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (li['item_id'], invoice_id, li['transaction_id'],
                      li['description'], li['sku'], li['qty'],
                      li['unit_cost'], li['markup_pct'], li['unit_price'],
                      li['line_total'], li['sort_order']))

            # Mark transactions as invoiced
            cur.execute("""
                UPDATE transactions SET fulfillment_status='invoiced',
                    fulfillment_status_updated_at=NOW()
                WHERE transaction_id = ANY(%s::uuid[])
            """, (txn_ids,))

        flash(f'Invoice {inv_number} created!', 'success')
        return redirect(url_for('invoicing.view_invoice', invoice_id=invoice_id))

    # GET — load received orders
    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT t.transaction_id, t.order_number, t.retailer, t.purchase_date,
                   t.price_total, t.fulfillment_status,
                   u.username AS person_name, c.company_name, c.company_id
            FROM transactions t
            LEFT JOIN dim_users u      ON t.user_id = u.user_id
            LEFT JOIN dim_companies c  ON t.company_id = c.company_id
            WHERE t.fulfillment_status='received' AND t.is_active=TRUE
            ORDER BY t.retailer, t.order_number
        """)
        orders = cur.fetchall()

        cur.execute("SELECT company_id, company_name, company_short_code FROM dim_companies WHERE is_active=TRUE ORDER BY company_name")
        companies = cur.fetchall()
        cur.execute("SELECT customer_id, customer_name FROM dim_customers WHERE is_active=TRUE ORDER BY customer_name")
        customers = cur.fetchall()

    return render_template('invoicing/new.html',
                           orders=orders, companies=companies, customers=customers,
                           today=str(date.today()))


# ── Invoice Detail ────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>')
@login_required
@require_role('admin')
def view_invoice(invoice_id):
    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code,
                   cu.customer_name, u.username AS created_by_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            LEFT JOIN dim_users u      ON i.created_by  = u.user_id
            WHERE i.invoice_id=%s
        """, (str(invoice_id),))
        inv = cur.fetchone()
        if not inv:
            flash('Invoice not found.', 'error')
            return redirect(url_for('invoicing.index'))

        cur.execute("""
            SELECT * FROM invoice_items WHERE invoice_id=%s ORDER BY sort_order
        """, (str(invoice_id),))
        items = cur.fetchall()

    code = (inv['company_short_code'] or 'SE').upper()
    co   = COMPANY_DATA.get(code, COMPANY_DATA['SE'])
    return render_template('invoicing/view.html', inv=inv, items=items, co=co)


# ── Update Status ─────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/status', methods=['POST'])
@login_required
@require_role('admin')
def update_status(invoice_id):
    status = request.form.get('status')
    if status in ('draft', 'sent', 'paid'):
        with db_cursor() as (cur, conn):
            cur.execute("UPDATE invoices SET status=%s WHERE invoice_id=%s",
                        (status, str(invoice_id)))
    return redirect(url_for('invoicing.view_invoice', invoice_id=str(invoice_id)))


# ── Export Excel ──────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/export-excel')
@login_required
@require_role('admin')
def export_excel(invoice_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code,
                   cu.customer_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            WHERE i.invoice_id=%s
        """, (str(invoice_id),))
        inv = cur.fetchone()
        cur.execute("SELECT * FROM invoice_items WHERE invoice_id=%s ORDER BY sort_order",
                    (str(invoice_id),))
        items = cur.fetchall()

    code = (inv['company_short_code'] or 'SE').upper()
    co   = COMPANY_DATA.get(code, COMPANY_DATA['SE'])

    wb = Workbook()
    ws = wb.active
    ws.title = inv['invoice_number']

    # Column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    thin = Side(style='thin')
    border = Border(bottom=Side(style='medium'))

    def cell(row, col, value, bold=False, align='left', fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, name='Calibri', size=11)
        c.alignment = Alignment(horizontal=align, vertical='center')
        if fmt: c.number_format = fmt
        return c

    # Header
    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value=co['name'])
    c.font = Font(bold=True, size=14, name='Calibri')
    ws.row_dimensions[1].height = 24

    addr_lines = co['addr'].split('\n')
    for i, line in enumerate(addr_lines):
        ws.merge_cells(f'A{2+i}:D{2+i}')
        cell(2+i, 1, line)
    r = 2 + len(addr_lines)

    ws.merge_cells(f'A{r}:B{r}')
    cell(r, 1, f"INVOICE #{inv['invoice_number']}", bold=True)
    cell(r, 3, 'Date:', bold=True, align='right')
    dt = inv['invoice_date']
    cell(r, 4, dt.strftime('%m/%d/%y') if hasattr(dt,'strftime') else str(dt), align='right')
    r += 1

    # To section
    r += 1
    cell(r, 1, 'To', bold=True)
    r += 1
    cell(r, 1, inv['customer_name'] or '—', bold=True)
    r += 2

    # Items header
    ws.row_dimensions[r].height = 18
    for col, (txt, al) in enumerate([('Description','left'),('Price','right'),('QTY','right'),('Total','right')], 1):
        c = ws.cell(row=r, column=col, value=txt)
        c.font = Font(bold=True, name='Calibri', size=11)
        c.alignment = Alignment(horizontal=al)
        c.fill = PatternFill('solid', fgColor='EEEEEE')
    r += 1

    for item in items:
        cell(r, 1, item['item_description'])
        cell(r, 2, float(item['unit_price']), align='right', fmt='"$"#,##0.00')
        cell(r, 3, int(item['quantity']), align='right')
        cell(r, 4, float(item['line_total']), align='right', fmt='"$"#,##0.00')
        r += 1

    r += 1
    cell(r, 3, 'SUBTOTAL', bold=True, align='right')
    cell(r, 4, float(inv['subtotal']), align='right', fmt='"$"#,##0.00')
    r += 1
    cell(r, 3, 'TAX RATE', align='right')
    cell(r, 4, '0.00%', align='right')
    r += 1
    other_lbl = inv.get('other_label') or 'OTHER'
    cell(r, 3, other_lbl, align='right')
    cell(r, 4, float(inv['other_amount'] or 0), align='right', fmt='"$"#,##0.00')
    r += 1
    cell(r, 3, 'TOTAL', bold=True, align='right')
    cell(r, 4, float(inv['total']), bold=True, align='right', fmt='"$"#,##0.00')
    r += 2

    cell(r, 1, 'THANK YOU FOR YOUR BUSINESS!', bold=True)
    r += 2

    cell(r, 1, 'TERMS & CONDITIONS', bold=True)
    r += 1
    cell(r, 1, 'Beneficiario:')
    r += 1
    for line in co['beneficiario'].split('\n'):
        cell(r, 1, line); r += 1
    r += 1
    cell(r, 1, 'Banco:')
    r += 1
    cell(r, 1, co['banco']); r += 1
    cell(r, 1, f"Route number: {co['route']}"); r += 1
    cell(r, 1, f"Account number: {co['account']}"); r += 2
    cell(r, 1, f"Make all checks payable to {co['payable']}")

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Invoice_{inv['invoice_number']}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ── Export PDF ────────────────────────────────────────────────────────────────

@invoicing_bp.route('/<uuid:invoice_id>/export-pdf')
@login_required
@require_role('admin')
def export_pdf(invoice_id):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

    with db_cursor() as (cur, _):
        cur.execute("""
            SELECT i.*, c.company_name, c.company_short_code,
                   cu.customer_name
            FROM invoices i
            LEFT JOIN dim_companies c  ON i.company_id  = c.company_id
            LEFT JOIN dim_customers cu ON i.customer_id = cu.customer_id
            WHERE i.invoice_id=%s
        """, (str(invoice_id),))
        inv = cur.fetchone()
        cur.execute("SELECT * FROM invoice_items WHERE invoice_id=%s ORDER BY sort_order",
                    (str(invoice_id),))
        items = cur.fetchall()

    code = (inv['company_short_code'] or 'SE').upper()
    co   = COMPANY_DATA.get(code, COMPANY_DATA['SE'])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    normal  = ParagraphStyle('n', fontName='Helvetica', fontSize=10, leading=14)
    bold    = ParagraphStyle('b', fontName='Helvetica-Bold', fontSize=10, leading=14)
    right   = ParagraphStyle('r', fontName='Helvetica', fontSize=10, leading=14, alignment=TA_RIGHT)
    rbold   = ParagraphStyle('rb', fontName='Helvetica-Bold', fontSize=10, leading=14, alignment=TA_RIGHT)
    center  = ParagraphStyle('c', fontName='Helvetica', fontSize=10, leading=14, alignment=TA_CENTER)
    title   = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=16, leading=20)
    small   = ParagraphStyle('s', fontName='Helvetica', fontSize=9, leading=12, textColor=colors.HexColor('#666666'))

    story = []

    # Header
    addr_html = co['addr'].replace('\n','<br/>')
    inv_date  = inv['invoice_date']
    date_str  = inv_date.strftime('%m/%d/%y') if hasattr(inv_date,'strftime') else str(inv_date)
    header_data = [
        [Paragraph(co['name'], title),
         Paragraph(f"<b>INVOICE #{inv['invoice_number']}</b>", rbold)],
        [Paragraph(addr_html, small),
         Paragraph(date_str, right)],
    ]
    ht = Table(header_data, colWidths=[4*inch, 3*inch])
    ht.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(ht)
    story.append(Spacer(1, 0.2*inch))

    # To
    story.append(Paragraph('To', small))
    story.append(Paragraph(inv['customer_name'] or '—', bold))
    story.append(Spacer(1, 0.25*inch))

    # Line items table
    table_data = [[
        Paragraph('Description', bold),
        Paragraph('Price', rbold),
        Paragraph('QTY', rbold),
        Paragraph('Total', rbold),
    ]]
    for item in items:
        table_data.append([
            Paragraph(item['item_description'] or '', normal),
            Paragraph(f"${float(item['unit_price']):,.2f}", right),
            Paragraph(str(item['quantity']), right),
            Paragraph(f"${float(item['line_total']):,.2f}", right),
        ])

    col_w = [3.8*inch, 1.2*inch, 0.7*inch, 1.3*inch]
    t = Table(table_data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#EEEEEE')),
        ('LINEBELOW',     (0,0), (-1,0), 0.5, colors.black),
        ('LINEBELOW',     (0,1), (-1,-1), 0.25, colors.HexColor('#DDDDDD')),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.15*inch))

    # Totals
    other_lbl = inv.get('other_label') or 'OTHER'
    totals_data = [
        [Paragraph('SUBTOTAL', right),   Paragraph(f"${float(inv['subtotal']):,.2f}", right)],
        [Paragraph('TAX RATE', right),   Paragraph('0.00%', right)],
        [Paragraph(other_lbl, right),    Paragraph(f"${float(inv['other_amount'] or 0):,.2f}", right)],
        [Paragraph('<b>TOTAL</b>', rbold), Paragraph(f"<b>${float(inv['total']):,.2f}</b>", rbold)],
    ]
    tt = Table(totals_data, colWidths=[5.5*inch, 1.5*inch])
    tt.setStyle(TableStyle([
        ('LINEABOVE',     (0,3), (-1,3), 0.5, colors.black),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(tt)
    story.append(Spacer(1, 0.3*inch))

    story.append(Paragraph('THANK YOU FOR YOUR BUSINESS!', center))
    story.append(Spacer(1, 0.25*inch))

    # Banking
    bene_html = co['beneficiario'].replace('\n','<br/>')
    story.append(Paragraph('<b>TERMS &amp; CONDITIONS</b>', bold))
    story.append(Spacer(1, 0.08*inch))
    banking_data = [
        [Paragraph('<b>Beneficiario:</b>', bold), Paragraph('<b>Banco:</b>', bold)],
        [Paragraph(bene_html, normal),
         Paragraph(f"{co['banco']}<br/>Route number: {co['route']}<br/>Account number: {co['account']}", normal)],
    ]
    bt = Table(banking_data, colWidths=[3.5*inch, 3.5*inch])
    bt.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(bt)
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(f"Make all checks payable to {co['payable']}", small))

    doc.build(story)
    buf.seek(0)
    fname = f"Invoice_{inv['invoice_number']}.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname)
